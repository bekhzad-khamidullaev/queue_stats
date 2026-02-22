from __future__ import annotations

import mimetypes
import os
import re
import time
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List
from urllib.parse import urlencode
from urllib.parse import urlparse, urlunparse

import requests
from django.conf import settings as django_settings
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.db import connections
from django.http import FileResponse, Http404, HttpRequest, HttpResponse, StreamingHttpResponse
from django.shortcuts import redirect, render
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas

from accounts.models import UserRoles
from settings.models import AgentDisplayMapping, GeneralSettings, OperatorPayoutRate, QueueDisplayMapping
from .ami_manager import AMIManager
from .i18n_map import tr as i18n_tr
from .models import CallTranscription
from .views import _fetch_queuelog_rows, _normalize_list, _parse_datetime

_AGENT_EXT_RE = re.compile(r"\b([0-9]{3,6})\b")
_FILTER_SESSION_KEY = "ui_saved_filters"
_LIST_FILTER_KEYS = {"queues", "agents"}
_SCALAR_FILTER_KEYS = {
    "start",
    "end",
    "src",
    "dst",
    "disposition",
    "channel",
    "caller",
    "q",
    "page_size",
}
_PERSISTED_FILTER_KEYS = _LIST_FILTER_KEYS | _SCALAR_FILTER_KEYS


def _user_allowed(request: HttpRequest) -> bool:
    return request.user.is_authenticated and request.user.role in {
        UserRoles.ADMIN,
        UserRoles.SUPERVISOR,
        UserRoles.ANALYST,
        UserRoles.AGENT,
    }


def _admin_allowed(request: HttpRequest) -> bool:
    return request.user.is_authenticated and request.user.role == UserRoles.ADMIN


def _saved_filters(request: HttpRequest) -> Dict[str, Any]:
    raw = request.session.get(_FILTER_SESSION_KEY, {})
    return raw if isinstance(raw, dict) else {}


def _parse_param_list(raw_items: List[str], raw_single: str | None = None) -> List[str]:
    parsed: List[str] = []
    if raw_items:
        for item in raw_items:
            parsed.extend(_normalize_list(item))
    else:
        parsed.extend(_normalize_list(raw_single))

    unique: List[str] = []
    for item in parsed:
        if item not in unique:
            unique.append(item)
    return unique


def _persist_filters(request: HttpRequest) -> None:
    if request.method != "GET":
        return
    if not request.GET:
        return

    current = dict(_saved_filters(request))
    updated = False
    for key in _PERSISTED_FILTER_KEYS:
        if key not in request.GET:
            continue
        if key in _LIST_FILTER_KEYS:
            value: Any = _parse_param_list(request.GET.getlist(key), request.GET.get(key))
        else:
            value = (request.GET.get(key) or "").strip()
        if current.get(key) != value:
            current[key] = value
            updated = True
    if updated:
        request.session[_FILTER_SESSION_KEY] = current
        request.session.modified = True


def _filter_value(request: HttpRequest, key: str, default: str = "") -> str:
    if key in request.GET:
        return (request.GET.get(key) or "").strip()
    saved = _saved_filters(request)
    value = saved.get(key, default)
    return str(value).strip() if value is not None else default


def _filter_list(request: HttpRequest, key: str) -> List[str]:
    if key in request.GET:
        return _parse_param_list(request.GET.getlist(key), request.GET.get(key))
    saved = _saved_filters(request)
    saved_value = saved.get(key, [])
    if isinstance(saved_value, list):
        raw_items = [str(item) for item in saved_value]
        return _parse_param_list(raw_items)
    return _parse_param_list([], str(saved_value))


def _interval_from_request(request: HttpRequest) -> tuple[datetime, datetime]:
    start = _parse_datetime(_filter_value(request, "start"))
    end = _parse_datetime(_filter_value(request, "end"))
    if not end:
        end = datetime.now().replace(hour=23, minute=59, second=59)
    if not start:
        start = end.replace(hour=0, minute=0, second=0)
    return start, end


def _get_param_list(request: HttpRequest, key: str) -> List[str]:
    return _filter_list(request, key)


def _queue_map() -> Dict[str, str]:
    return {item.queue_system_name: item.queue_display_name for item in QueueDisplayMapping.objects.all()}


def _agent_aliases(value: str) -> List[str]:
    raw = str(value or "").strip()
    if not raw:
        return []
    aliases: List[str] = [raw]

    current = raw
    if "/" in current:
        current = current.split("/", 1)[1].strip()
        aliases.append(current)
    if "@" in current:
        current = current.split("@", 1)[0].strip()
        aliases.append(current)
    if ";" in current:
        current = current.split(";", 1)[0].strip()
        aliases.append(current)
    if "-" in current and current.rsplit("-", 1)[1].isdigit():
        aliases.append(current.rsplit("-", 1)[0].strip())

    for match in _AGENT_EXT_RE.findall(raw):
        aliases.append(match)

    unique: List[str] = []
    for item in aliases:
        if item and item not in unique:
            unique.append(item)
    return unique


def _agent_map() -> Dict[str, str]:
    amap: Dict[str, str] = {}
    for item in AgentDisplayMapping.objects.all():
        display = item.agent_display_name
        for key in _agent_aliases(item.agent_system_name):
            amap.setdefault(key, display)
    return amap


def _is_valid_agent_value(value: str) -> bool:
    v = (value or "").strip()
    if len(v) < 2:
        return False
    return any(ch.isdigit() for ch in v) or any(ch.isalpha() for ch in v)


def _display_queue(value: str, qmap: Dict[str, str]) -> str:
    return qmap.get(value, value)


def _display_agent(value: str, amap: Dict[str, str]) -> str:
    for key in _agent_aliases(value):
        if key in amap:
            return amap[key]
    aliases = _agent_aliases(value)
    return aliases[-1] if aliases else value


def _human_party(value: str, amap: Dict[str, str]) -> str:
    raw = str(value or "").strip()
    if not raw or raw.lower() in {"<unknown>", "unknown"}:
        return i18n_tr("Неизвестно")
    display = _display_agent(raw, amap)
    aliases = _agent_aliases(raw)
    ext = next((a for a in aliases if a.isdigit() and 2 <= len(a) <= 6), "")
    if ext and display != ext:
        return f"{display} ({ext})"
    return display


def _human_channel(value: str, amap: Dict[str, str]) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    tech = raw.split("/", 1)[0] if "/" in raw else ""
    base = raw
    if "-" in base:
        left, right = base.rsplit("-", 1)
        if right and all(ch in "0123456789abcdefABCDEF" for ch in right):
            base = left
    display = _display_agent(base, amap)
    return f"{display} [{tech}]" if tech and display != base else raw


def _extract_operator_ext(value: str) -> str:
    for alias in _agent_aliases(value):
        if alias.isdigit() and 2 <= len(alias) <= 6:
            return alias
    return ""


def _duration_to_seconds(value: Any) -> int:
    raw = str(value or "").strip()
    if not raw:
        return 0
    if raw.isdigit():
        return int(raw)
    parts = raw.split(":")
    if len(parts) != 3:
        return 0
    try:
        hours = int(parts[0])
        minutes = int(parts[1])
        seconds = int(parts[2])
    except (TypeError, ValueError):
        return 0
    return (hours * 3600) + (minutes * 60) + seconds


def _channel_row_rank(row: Dict[str, Any]) -> tuple[int, int, int]:
    app = str(row.get("Application", "")).strip().lower()
    app_priority = {
        "queue": 4,
        "dial": 3,
        "bridge": 3,
        "appqueue": 2,
    }.get(app, 1)
    caller = str(row.get("CallerIDNum", "")).strip()
    connected = str(row.get("ConnectedLineNum", "")).strip()
    known_parties = int(bool(caller and caller.lower() not in {"<unknown>", "unknown"})) + int(
        bool(connected and connected.lower() not in {"<unknown>", "unknown"})
    )
    duration = _duration_to_seconds(row.get("Duration", ""))
    return app_priority, known_parties, duration


def _payout_rate_map() -> Dict[str, float]:
    rates: Dict[str, Decimal] = {}
    for item in OperatorPayoutRate.objects.all():
        for key in _agent_aliases(item.agent_system_name):
            rates.setdefault(key, Decimal(item.rate_per_minute))
    return rates


def _unique_non_empty(values: List[str]) -> List[str]:
    seen = set()
    unique: List[str] = []
    for value in values:
        item = str(value or "").strip()
        if item and item not in seen:
            seen.add(item)
            unique.append(item)
    return unique


def _is_internal_party(value: str) -> bool:
    raw = str(value or "").strip()
    if not raw:
        return False
    aliases = _agent_aliases(raw)
    if any(alias.isdigit() and 2 <= len(alias) <= 6 for alias in aliases):
        return True
    lowered = raw.lower()
    return lowered.startswith(("sip/", "pjsip/", "local/", "agent/"))


def _classify_call_direction(src: str, dst: str, dcontext: str, queue_agent: str) -> str:
    src_internal = _is_internal_party(src)
    dst_internal = _is_internal_party(dst)
    dcontext_lower = str(dcontext or "").strip().lower()

    if src_internal and not dst_internal:
        return "outgoing"
    if dst_internal and not src_internal:
        return "incoming"
    if queue_agent:
        return "incoming"
    if "from-internal" in dcontext_lower:
        return "outgoing"
    if any(token in dcontext_lower for token in ("from-trunk", "from-pstn", "from-external")):
        return "incoming"
    return "unknown"


def _get_available_queues() -> List[str]:
    queues: List[str] = []
    with connections["default"].cursor() as cursor:
        try:
            cursor.execute("SELECT DISTINCT queuename FROM queues_new WHERE queuename IS NOT NULL AND queuename <> '' ORDER BY queuename")
            queues.extend([str(row[0]).strip() for row in cursor.fetchall()])
        except Exception:
            pass
        if not _unique_non_empty(queues):
            try:
                cursor.execute("SELECT DISTINCT queuename FROM queuelog WHERE queuename IS NOT NULL AND queuename <> '' ORDER BY queuename")
                queues.extend([str(row[0]).strip() for row in cursor.fetchall()])
            except Exception:
                pass

    queues.extend(list(QueueDisplayMapping.objects.values_list("queue_system_name", flat=True)))
    return _unique_non_empty(queues)


def _get_available_agents() -> List[Dict[str, str]]:
    out: List[Dict[str, str]] = []
    seen = set()

    def _add_agent(agent_value: str, name_value: str = "") -> None:
        agent = str(agent_value or "").strip()
        if (not agent) or (not _is_valid_agent_value(agent)) or agent in seen:
            return
        seen.add(agent)
        out.append({"agent": agent, "name": str(name_value or "").strip()})

    with connections["default"].cursor() as cursor:
        try:
            cursor.execute(
                """
                SELECT agent, MAX(COALESCE(name, '')) AS name
                FROM agents_new
                WHERE agent IS NOT NULL AND agent <> ''
                GROUP BY agent
                ORDER BY agent
                """
            )
            for row in cursor.fetchall():
                _add_agent(str(row[0]), str(row[1] or ""))
        except Exception:
            pass

        if not out:
            try:
                cursor.execute(
                    """
                    SELECT agent, '' AS name
                    FROM queuelog
                    WHERE agent IS NOT NULL AND agent <> ''
                    GROUP BY agent
                    ORDER BY agent
                    """
                )
                for row in cursor.fetchall():
                    _add_agent(str(row[0]), "")
            except Exception:
                pass

    for system_name in AgentDisplayMapping.objects.values_list("agent_system_name", flat=True):
        _add_agent(str(system_name), "")
    return out


def _caller_map_by_callids(callids: List[str]) -> Dict[str, str]:
    clean = [c for c in callids if c]
    if not clean:
        return {}
    out: Dict[str, str] = {}
    chunk_size = 500
    with connections["default"].cursor() as cursor:
        for i in range(0, len(clean), chunk_size):
            chunk = clean[i:i + chunk_size]
            placeholders = ",".join(["%s"] * len(chunk))
            cursor.execute(
                f"""
                SELECT callid,
                       SUBSTRING_INDEX(
                         GROUP_CONCAT(NULLIF(data2, '') ORDER BY time DESC SEPARATOR ','),
                         ',',
                         1
                       ) AS caller
                FROM queuelog
                WHERE event = 'ENTERQUEUE' AND callid IN ({placeholders})
                GROUP BY callid
                """,
                chunk,
            )
            for callid, caller in cursor.fetchall():
                out[str(callid)] = str(caller or "")
    return out


def _base_context(request: HttpRequest) -> Dict[str, Any]:
    _persist_filters(request)
    now = datetime.now()
    selected_queues_list = _get_param_list(request, "queues")
    selected_agents_list = _get_param_list(request, "agents")
    start_value = _filter_value(request, "start")
    end_value = _filter_value(request, "end")
    queues = _get_available_queues()
    agents = _get_available_agents()
    qmap = _queue_map()
    amap = _agent_map()

    queue_options = [{"value": q, "label": _display_queue(q, qmap)} for q in queues]
    agent_options: List[Dict[str, str]] = []
    for item in agents:
        agent_system = str(item.get("agent") or "")
        agent_name = str(item.get("name") or "").strip()
        agent_display = _display_agent(agent_system, amap)
        if agent_display != agent_system:
            label = f"{agent_display} ({agent_system})"
        elif agent_name and agent_name != agent_system:
            label = f"{agent_system} ({agent_name})"
        else:
            label = agent_system
        agent_options.append({"value": agent_system, "label": label})

    return {
        "queues": queues,
        "agents": agents,
        "queue_options": queue_options,
        "agent_options": agent_options,
        "default_start": start_value or now.replace(hour=0, minute=0, second=0).strftime("%Y-%m-%d %H:%M:%S"),
        "default_end": end_value or now.replace(hour=23, minute=59, second=59).strftime("%Y-%m-%d %H:%M:%S"),
        "selected_queues": ",".join(selected_queues_list),
        "selected_agents": ",".join(selected_agents_list),
        "selected_queues_list": selected_queues_list,
        "selected_agents_list": selected_agents_list,
        "selected_src": _filter_value(request, "src"),
        "selected_dst": _filter_value(request, "dst"),
        "selected_disposition": _filter_value(request, "disposition"),
        "selected_channel": _filter_value(request, "channel"),
        "selected_caller": _filter_value(request, "caller"),
        "selected_query": _filter_value(request, "q"),
        "selected_page_size": _filter_value(request, "page_size", "100"),
    }


def _get_general_settings() -> GeneralSettings:
    obj = GeneralSettings.objects.first()
    if obj:
        return obj
    return GeneralSettings.objects.create()


def _to_int(value: Any, default: int, minimum: int | None = None, maximum: int | None = None) -> int:
    try:
        parsed = int(str(value))
    except (TypeError, ValueError):
        parsed = default
    if minimum is not None:
        parsed = max(minimum, parsed)
    if maximum is not None:
        parsed = min(maximum, parsed)
    return parsed


def _paginated_rows(request: HttpRequest, rows: List[Dict[str, Any]], default_page_size: int = 100, max_page_size: int = 500) -> tuple[List[Dict[str, Any]], Dict[str, Any]]:
    page = _to_int(request.GET.get("page"), default=1, minimum=1)
    page_size = _to_int(request.GET.get("page_size"), default=default_page_size, minimum=1, maximum=max_page_size)

    total = len(rows)
    total_pages = max(1, (total + page_size - 1) // page_size) if page_size else 1
    if page > total_pages:
        page = total_pages
    offset = (page - 1) * page_size
    page_rows = rows[offset:offset + page_size]

    start_index = offset + 1 if total else 0
    end_index = min(offset + page_size, total)
    query_pairs = [(key, value) for key, value in request.GET.lists() if key != "page"]
    base_qs = urlencode(query_pairs, doseq=True)

    def _link(target_page: int) -> str:
        if base_qs:
            return f"?{base_qs}&page={target_page}"
        return f"?page={target_page}"

    return page_rows, {
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "start_index": start_index,
        "end_index": end_index,
        "has_prev": page > 1,
        "has_next": page < total_pages,
        "prev_page_url": _link(page - 1) if page > 1 else "",
        "next_page_url": _link(page + 1) if page < total_pages else "",
    }


def _pagination_params(request: HttpRequest, default_page_size: int = 100, max_page_size: int = 500) -> tuple[int, int]:
    page = _to_int(request.GET.get("page"), default=1, minimum=1)
    page_size = _to_int(_filter_value(request, "page_size", str(default_page_size)), default=default_page_size, minimum=1, maximum=max_page_size)
    return page, page_size


def _pagination_meta(request: HttpRequest, page: int, page_size: int, total: int) -> Dict[str, Any]:
    total_pages = max(1, (total + page_size - 1) // page_size) if page_size else 1
    if page > total_pages:
        page = total_pages
    offset = (page - 1) * page_size
    start_index = offset + 1 if total else 0
    end_index = min(offset + page_size, total)
    query_pairs = [(key, value) for key, value in request.GET.lists() if key != "page"]
    base_qs = urlencode(query_pairs, doseq=True)

    def _link(target_page: int) -> str:
        if base_qs:
            return f"?{base_qs}&page={target_page}"
        return f"?page={target_page}"

    return {
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "start_index": start_index,
        "end_index": end_index,
        "has_prev": page > 1,
        "has_next": page < total_pages,
        "prev_page_url": _link(page - 1) if page > 1 else "",
        "next_page_url": _link(page + 1) if page < total_pages else "",
    }


def _build_queuelog_filter_sql(
    start: datetime,
    end: datetime,
    queues: List[str] | None,
    agents: List[str] | None,
    events: List[str],
) -> tuple[str, List[Any]]:
    where = [
        "time >= %s",
        "time <= %s",
        f"event IN ({','.join(['%s'] * len(events))})",
    ]
    params: List[Any] = [start, end, *events]

    queue_params = list(queues or [])
    if queue_params:
        where.append(f"queuename IN ({','.join(['%s'] * len(queue_params))})")
        params.extend(queue_params)

    agent_params = list(agents or [])
    if agent_params:
        where.append(f"agent IN ({','.join(['%s'] * len(agent_params))})")
        params.extend(agent_params)

    return " AND ".join(where), params


def _fetch_queuelog_page(
    start: datetime,
    end: datetime,
    queues: List[str] | None,
    agents: List[str] | None,
    events: List[str],
    page: int,
    page_size: int,
) -> tuple[List[Dict[str, Any]], int]:
    where_sql, params = _build_queuelog_filter_sql(start, end, queues, agents, events)
    requested_page = max(1, page)
    safe_page_size = max(1, page_size)

    sql = f"""
        SELECT time, callid, queuename, agent, event, data1, data2, data3
        FROM queuelog
        WHERE {where_sql}
        ORDER BY time DESC
        LIMIT %s OFFSET %s
    """
    count_sql = f"SELECT COUNT(*) FROM queuelog WHERE {where_sql}"

    with connections["default"].cursor() as cursor:
        cursor.execute(count_sql, params)
        total = int(cursor.fetchone()[0] or 0)
        total_pages = max(1, (total + safe_page_size - 1) // safe_page_size) if safe_page_size else 1
        safe_page = min(requested_page, total_pages)
        offset = (safe_page - 1) * safe_page_size

        cursor.execute(sql, [*params, safe_page_size, offset])
        columns = [col[0] for col in cursor.description]
        rows = [dict(zip(columns, row)) for row in cursor.fetchall()]
    return rows, total


def _queuelog_avg_numeric(
    start: datetime,
    end: datetime,
    queues: List[str] | None,
    agents: List[str] | None,
    events: List[str],
    data_column: str,
) -> float:
    allowed_columns = {"data1", "data2", "data3"}
    if data_column not in allowed_columns:
        return 0.0
    where_sql, params = _build_queuelog_filter_sql(start, end, queues, agents, events)
    sql = f"SELECT COALESCE(AVG(CAST({data_column} AS UNSIGNED)), 0) FROM queuelog WHERE {where_sql}"
    with connections["default"].cursor() as cursor:
        cursor.execute(sql, params)
        avg_value = cursor.fetchone()[0]
    try:
        return round(float(avg_value or 0), 2)
    except (TypeError, ValueError):
        return 0.0


def _answered_dataset(request: HttpRequest) -> Dict[str, Any]:
    start, end = _interval_from_request(request)
    queues = _get_param_list(request, "queues")
    agents = _get_param_list(request, "agents")
    qmap = _queue_map()
    amap = _agent_map()

    page, page_size = _pagination_params(request)
    rows, total = _fetch_queuelog_page(start, end, queues, agents, ["COMPLETECALLER", "COMPLETEAGENT"], page, page_size)
    caller_map = _caller_map_by_callids([str(r.get("callid") or "") for r in rows])
    flat_rows: List[Dict[str, Any]] = []
    for row in rows:
        callid = str(row.get("callid") or "")
        queue = str(row.get("queuename") or "")
        agent = str(row.get("agent") or "")
        flat_rows.append(
            {
                "callid": callid,
                "time": row.get("time"),
                "caller": caller_map.get(callid, ""),
                "queue": queue,
                "queue_display": _display_queue(queue, qmap),
                "agent": agent,
                "agent_display": _display_agent(agent, amap),
                "hold": int(row.get("data1") or 0),
                "talk": int(row.get("data2") or 0),
            }
        )

    pagination = _pagination_meta(request, page, page_size, total)
    avg_hold = _queuelog_avg_numeric(start, end, queues, agents, ["COMPLETECALLER", "COMPLETEAGENT"], "data1")
    avg_talk = _queuelog_avg_numeric(start, end, queues, agents, ["COMPLETECALLER", "COMPLETEAGENT"], "data2")
    return {
        "start": start,
        "end": end,
        "rows": flat_rows,
        "total": total,
        "avg_hold": avg_hold,
        "avg_talk": avg_talk,
        **pagination,
    }


def _unanswered_dataset(request: HttpRequest) -> Dict[str, Any]:
    start, end = _interval_from_request(request)
    queues = _get_param_list(request, "queues")
    qmap = _queue_map()
    page, page_size = _pagination_params(request)
    rows, total = _fetch_queuelog_page(start, end, queues, None, ["ABANDON", "EXITWITHTIMEOUT"], page, page_size)
    caller_map = _caller_map_by_callids([str(r.get("callid") or "") for r in rows])

    flat_rows: List[Dict[str, Any]] = []
    for row in rows:
        callid = str(row.get("callid") or "")
        queue = str(row.get("queuename") or "")
        flat_rows.append(
            {
                "callid": callid,
                "time": row.get("time"),
                "caller": caller_map.get(callid, ""),
                "queue": queue,
                "queue_display": _display_queue(queue, qmap),
                "event": row.get("event"),
                "start_pos": int(row.get("data2") or 0),
                "end_pos": int(row.get("data1") or 0),
                "wait_sec": int(row.get("data3") or 0),
            }
        )

    pagination = _pagination_meta(request, page, page_size, total)
    avg_wait = _queuelog_avg_numeric(start, end, queues, None, ["ABANDON", "EXITWITHTIMEOUT"], "data3")
    return {
        "start": start,
        "end": end,
        "rows": flat_rows,
        "total": total,
        "avg_wait": avg_wait,
        **pagination,
    }


def _cdr_dataset(request: HttpRequest) -> Dict[str, Any]:
    start, end = _interval_from_request(request)
    agents = _get_param_list(request, "agents")
    src_filter = _filter_value(request, "src")
    dst_filter = _filter_value(request, "dst")
    disposition_filter = _filter_value(request, "disposition")
    page, page_size = _pagination_params(request)

    amap = _agent_map()
    qmap = _queue_map()

    where: List[str] = ["c.calldate >= %s", "c.calldate <= %s"]
    params: List[Any] = [start, end]

    agent_values: List[str] = []
    if agents:
        for raw in agents:
            for alias in _agent_aliases(raw):
                if alias not in agent_values:
                    agent_values.append(alias)
        if agent_values:
            placeholders = ",".join(["%s"] * len(agent_values))
            where.append(
                "("
                f"c.cnam IN ({placeholders}) OR c.cnum IN ({placeholders}) "
                f"OR ql.agent IN ({placeholders}) OR c.dstchannel IN ({placeholders})"
                ")"
            )
            params.extend(agent_values)
            params.extend(agent_values)
            params.extend(agent_values)
            params.extend(agent_values)

            like_chunks: List[str] = []
            like_params: List[str] = []
            for alias in agent_values:
                like_chunks.append("c.dstchannel LIKE %s")
                like_params.append(f"%/{alias}@%")
                like_chunks.append("c.dstchannel LIKE %s")
                like_params.append(f"%/{alias}-%")
            if like_chunks:
                where.append(f"({' OR '.join(like_chunks)})")
                params.extend(like_params)

    if src_filter:
        where.append("c.src LIKE %s")
        params.append(f"%{src_filter}%")
    if dst_filter:
        where.append("(c.dst LIKE %s OR c.dcontext LIKE %s OR c.lastdata LIKE %s)")
        params.append(f"%{dst_filter}%")
        params.append(f"%{dst_filter}%")
        params.append(f"%{dst_filter}%")
    if disposition_filter:
        where.append("c.disposition = %s")
        params.append(disposition_filter)

    where_sql = " AND ".join(where)
    from_sql = """
        FROM cdr c
        LEFT JOIN (
            SELECT
                callid,
                SUBSTRING_INDEX(GROUP_CONCAT(agent ORDER BY time DESC SEPARATOR ','), ',', 1) AS agent
            FROM queuelog
            WHERE event IN ('CONNECT', 'COMPLETECALLER', 'COMPLETEAGENT')
              AND time >= %s AND time <= %s
              AND agent IS NOT NULL AND agent <> ''
            GROUP BY callid
        ) ql ON ql.callid = c.uniqueid
    """
    count_sql = f"SELECT COUNT(*) {from_sql} WHERE {where_sql}"
    count_params = [start, end, *params]

    with connections["default"].cursor() as cursor:
        cursor.execute(count_sql, count_params)
        total = int(cursor.fetchone()[0] or 0)

    pagination = _pagination_meta(request, page, page_size, total)
    offset = (pagination["page"] - 1) * pagination["page_size"]

    sql = f"""
        SELECT
            c.uniqueid,
            c.calldate,
            c.src,
            c.dst,
            c.dcontext,
            c.dstchannel,
            c.lastapp,
            c.lastdata,
            c.cnum,
            c.cnam,
            c.duration,
            c.billsec,
            c.disposition,
            c.recordingfile,
            ql.agent AS queue_agent
        {from_sql}
        WHERE {where_sql}
        ORDER BY c.calldate DESC
        LIMIT %s OFFSET %s
    """
    data_params = [start, end, *params, pagination["page_size"], offset]

    with connections["default"].cursor() as cursor:
        cursor.execute(sql, data_params)
        rows = cursor.fetchall()
        columns = [col[0] for col in cursor.description]
    data = [dict(zip(columns, row)) for row in rows]

    for row in data:
        operator_system = str(row.get("queue_agent") or "")
        if not operator_system:
            operator_system = str(row.get("dstchannel") or "")
        row["operator_display"] = _display_agent(operator_system, amap)

        dst = str(row.get("dst") or "")
        if dst in {"", "s", "S"}:
            if str(row.get("lastapp") or "").lower() == "queue":
                queue_raw = str(row.get("lastdata") or "").split(",", 1)[0].strip()
                dst = _display_queue(queue_raw, qmap) if queue_raw else ""
            elif row.get("dcontext"):
                dst = _display_queue(str(row.get("dcontext") or ""), qmap)
        row["dst_display"] = dst or "-"

        row["has_recording"] = bool(row.get("recordingfile"))

    return {"start": start, "end": end, "rows": data, "total": total, **pagination}


def _call_detail_dataset(callid: str) -> Dict[str, Any]:
    qmap = _queue_map()
    amap = _agent_map()
    normalized_callid = str(callid or "").strip()
    cdr_row: Dict[str, Any] = {}

    with connections["default"].cursor() as cursor:
        cursor.execute(
            """
            SELECT
                uniqueid, calldate, src, dst, dcontext, channel, dstchannel,
                lastapp, lastdata, duration, billsec, disposition, recordingfile
            FROM cdr
            WHERE uniqueid = %s
            ORDER BY calldate DESC
            LIMIT 1
            """,
            [normalized_callid],
        )
        row = cursor.fetchone()
        if row:
            columns = [col[0] for col in cursor.description]
            cdr_row = dict(zip(columns, row))

        cursor.execute(
            """
            SELECT time, callid, queuename, agent, event, data1, data2, data3, data4, data5
            FROM queuelog
            WHERE callid = %s
            ORDER BY time ASC
            LIMIT 1000
            """,
            [normalized_callid],
        )
        qlog_columns = [col[0] for col in cursor.description]
        qlog_events = [dict(zip(qlog_columns, r)) for r in cursor.fetchall()]

    caller = ""
    event_rows: List[Dict[str, Any]] = []
    outcome = "IN_PROGRESS"
    for row in qlog_events:
        queue_raw = str(row.get("queuename") or "")
        agent_raw = str(row.get("agent") or "")
        event_name = str(row.get("event") or "")
        if event_name == "ENTERQUEUE" and not caller:
            caller = str(row.get("data2") or "")
        if event_name in {"COMPLETECALLER", "COMPLETEAGENT"}:
            outcome = "ANSWERED"
        elif event_name == "ABANDON":
            outcome = "ABANDONED"
        elif event_name == "EXITWITHTIMEOUT":
            outcome = "TIMEOUT"
        event_rows.append(
            {
                "time": row.get("time"),
                "event": event_name,
                "queue": queue_raw,
                "queue_display": _display_queue(queue_raw, qmap),
                "agent": agent_raw,
                "agent_display": _display_agent(agent_raw, amap) if agent_raw else "",
                "data1": row.get("data1"),
                "data2": row.get("data2"),
                "data3": row.get("data3"),
                "data4": row.get("data4"),
                "data5": row.get("data5"),
            }
        )

    if cdr_row:
        operator_system = str(cdr_row.get("dstchannel") or "")
        cdr_row["operator_display"] = _display_agent(operator_system, amap) if operator_system else ""
        cdr_row["has_recording"] = bool(cdr_row.get("recordingfile"))

    if not caller and cdr_row:
        caller = str(cdr_row.get("src") or "")

    transcription = CallTranscription.objects.filter(callid=normalized_callid).first()
    conf = _get_general_settings()
    transcription_configured = bool((conf.transcription_url or "").strip() and (conf.transcription_api_key or "").strip())
    transcription_text = str(transcription.text or "").strip() if transcription else ""
    transcription_chunks = [part.strip() for part in re.split(r"(?<=[.!?])\s+", transcription_text) if part.strip()]

    return {
        "callid": normalized_callid,
        "caller": caller,
        "outcome": outcome,
        "cdr": cdr_row,
        "events": event_rows,
        "has_data": bool(cdr_row) or bool(event_rows),
        "transcription": transcription,
        "transcription_text": transcription_text,
        "transcription_chunks": transcription_chunks,
        "transcription_configured": transcription_configured,
    }


def _fetch_recording_bytes_for_call(callid: str) -> tuple[str, bytes, str]:
    recording_path = _recording_file_by_uniqueid(callid)
    local_path = _resolve_recording_local_path(recording_path)
    if local_path:
        content_type = mimetypes.guess_type(str(local_path))[0] or "application/octet-stream"
        return local_path.name, local_path.read_bytes(), content_type

    conf = _get_general_settings()
    if not conf.download_url:
        raise Http404("Recording not available")

    params = {"url": recording_path, "token": conf.download_token}
    auth = (conf.download_user, conf.download_password)
    upstream = requests.get(conf.download_url, params=params, auth=auth, timeout=(5, 60))
    upstream.raise_for_status()
    filename = os.path.basename(recording_path) or f"{callid}.wav"
    content_type = upstream.headers.get("Content-Type", "application/octet-stream")
    return filename, upstream.content, content_type


def _transcribe_call(callid: str) -> tuple[bool, str]:
    conf = _get_general_settings()
    api_url = str(conf.transcription_url or "").strip()
    api_key = str(conf.transcription_api_key or "").strip()
    if not api_url or not api_key:
        return False, i18n_tr("Не настроен URL/API ключ сервиса транскрибации")

    transcription, _ = CallTranscription.objects.get_or_create(callid=callid)
    transcription.status = CallTranscription.Status.PROCESSING
    transcription.error_message = ""
    transcription.save(update_fields=["status", "error_message", "updated_at"])

    try:
        filename, file_bytes, content_type = _fetch_recording_bytes_for_call(callid)
        request_headers = {"X-API-KEY": api_key, "Accept": "application/json"}
        request_files = {"file": (filename, file_bytes, content_type)}
        request_timeout = (15, 900)
        request_attempts = 3

        candidate_urls = [api_url]
        parsed = urlparse(api_url)
        if parsed.scheme == "https":
            # Some endpoints are plain HTTP but mistakenly configured as HTTPS.
            candidate_urls.append(urlunparse(parsed._replace(scheme="http")))

        response = None
        last_error: requests.RequestException | None = None
        for target_url in candidate_urls:
            for attempt in range(1, request_attempts + 1):
                try:
                    response = requests.post(
                        target_url,
                        headers=request_headers,
                        files=request_files,
                        timeout=request_timeout,
                    )
                    response.raise_for_status()
                    break
                except (requests.Timeout, requests.ConnectionError) as exc:
                    last_error = exc
                    if attempt < request_attempts:
                        time.sleep(1)
                        continue
                    break
                except requests.RequestException as exc:
                    last_error = exc
                    break
            if response is not None:
                break

        if response is None:
            if last_error is not None:
                raise last_error
            raise requests.RequestException(i18n_tr("Не удалось получить ответ от сервиса транскрибации"))

        response.raise_for_status()
        payload = response.json() if response.content else {}
        text = str(payload.get("text") or "").strip()
        if not text:
            raise ValueError(i18n_tr("Сервис вернул пустой текст"))

        transcription.status = CallTranscription.Status.SUCCESS
        transcription.text = text
        transcription.error_message = ""
        transcription.save(update_fields=["status", "text", "error_message", "updated_at"])
        return True, i18n_tr("Транскрипт успешно получен")
    except Http404:
        error_message = i18n_tr("Аудиозапись для звонка не найдена")
    except requests.RequestException as exc:
        error_message = f"{i18n_tr('Ошибка запроса к сервису транскрибации')}: {exc}"
    except (ValueError, TypeError) as exc:
        error_message = str(exc)
    except Exception as exc:
        error_message = f"{i18n_tr('Ошибка транскрибации')}: {exc}"

    transcription.status = CallTranscription.Status.FAILED
    transcription.error_message = error_message
    transcription.save(update_fields=["status", "error_message", "updated_at"])
    return False, error_message


def _summary_dataset(request: HttpRequest) -> Dict[str, Any]:
    start, end = _interval_from_request(request)
    queues = _get_param_list(request, "queues")
    agents = _get_param_list(request, "agents")
    qmap = _queue_map()

    rows = _fetch_queuelog_rows(
        start,
        end,
        queues,
        agents,
        ["COMPLETECALLER", "COMPLETEAGENT", "ABANDON", "EXITWITHTIMEOUT"],
    )

    answered = 0
    abandoned = 0
    timeout = 0
    wait_total = 0
    talk_total = 0
    by_queue: Dict[str, Dict[str, int]] = {}

    for row in rows:
        queue = str(row.get("queuename") or "UNKNOWN")
        by_queue.setdefault(queue, {"answered": 0, "unanswered": 0, "total": 0})
        event = str(row.get("event") or "")
        if event in {"COMPLETECALLER", "COMPLETEAGENT"}:
            answered += 1
            by_queue[queue]["answered"] += 1
            wait_total += int(row.get("data1") or 0)
            talk_total += int(row.get("data2") or 0)
        elif event == "ABANDON":
            abandoned += 1
            by_queue[queue]["unanswered"] += 1
        elif event == "EXITWITHTIMEOUT":
            timeout += 1
            by_queue[queue]["unanswered"] += 1
        by_queue[queue]["total"] += 1

    total = answered + abandoned + timeout
    per_queue = []
    for system_name, values in sorted(by_queue.items(), key=lambda item: item[0]):
        per_queue.append({"queue": system_name, "queue_display": _display_queue(system_name, qmap), **values})

    return {
        "start": start,
        "end": end,
        "total": total,
        "answered": answered,
        "abandoned": abandoned,
        "timeout": timeout,
        "service_level": round(answered * 100 / total, 2) if total else 0,
        "avg_wait": round(wait_total / answered, 2) if answered else 0,
        "avg_talk": round(talk_total / answered, 2) if answered else 0,
        "per_queue": per_queue,
    }


def _analytics_dataset(request: HttpRequest) -> Dict[str, Any]:
    start, end = _interval_from_request(request)
    queues = _get_param_list(request, "queues")
    agents = _get_param_list(request, "agents")
    qmap = _queue_map()
    amap = _agent_map()
    queue_clause = ""
    agent_clause = ""
    if queues:
        queue_clause = f" AND queuename IN ({','.join(['%s'] * len(queues))})"
    if agents:
        agent_clause = f" AND agent IN ({','.join(['%s'] * len(agents))})"

    queue_params: List[Any] = [start, end, *queues]
    queue_agent_params: List[Any] = [start, end, *queues, *agents]

    sql_daily = f"""
        SELECT DATE(time) AS day,
               SUM(CASE WHEN event IN ('COMPLETECALLER','COMPLETEAGENT') THEN 1 ELSE 0 END) AS answered,
               SUM(CASE WHEN event = 'ABANDON' THEN 1 ELSE 0 END) AS abandoned,
               SUM(CASE WHEN event = 'EXITWITHTIMEOUT' THEN 1 ELSE 0 END) AS timeout,
               SUM(CASE WHEN event IN ('ABANDON','EXITWITHTIMEOUT') THEN 1 ELSE 0 END) AS unanswered
        FROM queuelog
        WHERE time >= %s AND time <= %s {queue_clause}
        GROUP BY day
        ORDER BY day
    """

    sql_hourly = f"""
        SELECT HOUR(time) AS hour,
               SUM(CASE WHEN event IN ('COMPLETECALLER','COMPLETEAGENT') THEN 1 ELSE 0 END) AS answered,
               SUM(CASE WHEN event IN ('ABANDON','EXITWITHTIMEOUT') THEN 1 ELSE 0 END) AS unanswered
        FROM queuelog
        WHERE time >= %s AND time <= %s {queue_clause}
        GROUP BY hour
        ORDER BY hour
    """

    sql_queue = f"""
        SELECT queuename,
               SUM(CASE WHEN event IN ('COMPLETECALLER','COMPLETEAGENT') THEN 1 ELSE 0 END) AS answered,
               SUM(CASE WHEN event = 'ABANDON' THEN 1 ELSE 0 END) AS abandoned,
               SUM(CASE WHEN event = 'EXITWITHTIMEOUT' THEN 1 ELSE 0 END) AS timeout,
               SUM(CASE WHEN event IN ('COMPLETECALLER','COMPLETEAGENT') THEN CAST(data1 AS UNSIGNED) ELSE 0 END) AS hold_sec,
               SUM(CASE WHEN event IN ('COMPLETECALLER','COMPLETEAGENT') THEN CAST(data2 AS UNSIGNED) ELSE 0 END) AS talk_sec,
               SUM(CASE WHEN event IN ('ABANDON','EXITWITHTIMEOUT') THEN 1 ELSE 0 END) AS unanswered
        FROM queuelog
        WHERE time >= %s AND time <= %s {queue_clause}
        GROUP BY queuename
        ORDER BY answered DESC
        LIMIT 200
    """

    sql_agents = f"""
        SELECT agent,
               COUNT(*) AS answered,
               COUNT(DISTINCT queuename) AS queues_count,
               SUM(CAST(data1 AS UNSIGNED)) AS hold_sec,
               SUM(CAST(data2 AS UNSIGNED)) AS talk_sec
        FROM queuelog
        WHERE time >= %s AND time <= %s {queue_clause}
          AND event IN ('COMPLETECALLER','COMPLETEAGENT') {agent_clause}
          AND agent IS NOT NULL AND agent <> ''
        GROUP BY agent
        ORDER BY answered DESC
        LIMIT 200
    """

    queue_clause_q = f" AND q.queuename IN ({','.join(['%s'] * len(queues))})" if queues else ""
    caller_agent_clause = (
        f"""
        AND EXISTS (
            SELECT 1
            FROM queuelog qa
            WHERE qa.callid = q.callid
              AND qa.event IN ('COMPLETECALLER','COMPLETEAGENT')
              AND qa.agent IN ({','.join(['%s'] * len(agents))})
        )
        """
        if agents
        else ""
    )
    caller_params: List[Any] = [start, end, *queues, *agents]
    sql_top_callers = f"""
        SELECT TRIM(q.data2) AS caller, COUNT(DISTINCT q.callid) AS calls
        FROM queuelog q
        WHERE q.time >= %s AND q.time <= %s
          AND q.event = 'ENTERQUEUE'
          {queue_clause_q}
          AND q.data2 IS NOT NULL
          AND TRIM(q.data2) <> ''
          AND LOWER(TRIM(q.data2)) NOT IN ('unknown','<unknown>')
          {caller_agent_clause}
        GROUP BY caller
        ORDER BY calls DESC, caller ASC
        LIMIT 25
    """

    cdr_queue_clause = (
        f"""
        AND EXISTS (
            SELECT 1
            FROM queuelog qf
            WHERE qf.callid = c.uniqueid
              AND qf.queuename IN ({','.join(['%s'] * len(queues))})
        )
        """
        if queues
        else ""
    )
    cdr_params: List[Any] = [start, end, *queues]
    sql_operator_cdr = f"""
        SELECT
            c.uniqueid,
            c.calldate,
            c.src,
            c.dst,
            c.dcontext,
            c.channel,
            c.dstchannel,
            c.billsec,
            ql.agent AS queue_agent
        FROM cdr c
        LEFT JOIN (
            SELECT
                callid,
                SUBSTRING_INDEX(GROUP_CONCAT(agent ORDER BY time DESC SEPARATOR ','), ',', 1) AS agent
            FROM queuelog
            WHERE event IN ('CONNECT', 'COMPLETECALLER', 'COMPLETEAGENT')
              AND agent IS NOT NULL AND agent <> ''
            GROUP BY callid
        ) ql ON ql.callid = c.uniqueid
        WHERE c.calldate >= %s AND c.calldate <= %s
          {cdr_queue_clause}
        ORDER BY c.calldate ASC
        LIMIT 50000
    """

    with connections["default"].cursor() as cursor:
        cursor.execute(sql_daily, queue_params)
        daily = [dict(zip([c[0] for c in cursor.description], row)) for row in cursor.fetchall()]

        cursor.execute(sql_hourly, queue_params)
        hourly = [dict(zip([c[0] for c in cursor.description], row)) for row in cursor.fetchall()]

        cursor.execute(sql_queue, queue_params)
        per_queue = [dict(zip([c[0] for c in cursor.description], row)) for row in cursor.fetchall()]

        cursor.execute(sql_agents, queue_agent_params)
        top_agents = [dict(zip([c[0] for c in cursor.description], row)) for row in cursor.fetchall()]

        cursor.execute(sql_top_callers, caller_params)
        top_callers = [dict(zip([c[0] for c in cursor.description], row)) for row in cursor.fetchall()]

        cursor.execute(sql_operator_cdr, cdr_params)
        operator_cdr_rows = [dict(zip([c[0] for c in cursor.description], row)) for row in cursor.fetchall()]

    total_answered = sum(int(x.get("answered") or 0) for x in daily)
    total_abandoned = sum(int(x.get("abandoned") or 0) for x in daily)
    total_timeout = sum(int(x.get("timeout") or 0) for x in daily)
    total_unanswered = total_abandoned + total_timeout
    total_calls = total_answered + total_unanswered

    sum_hold_answered = sum(int(x.get("hold_sec") or 0) for x in per_queue)
    sum_talk_answered = sum(int(x.get("talk_sec") or 0) for x in per_queue)

    for row in per_queue:
        answered = int(row.get("answered") or 0)
        abandoned = int(row.get("abandoned") or 0)
        timeout = int(row.get("timeout") or 0)
        unanswered = int(row.get("unanswered") or 0)
        total = answered + unanswered
        hold_sec = int(row.get("hold_sec") or 0)
        talk_sec = int(row.get("talk_sec") or 0)

        row["queue_display"] = _display_queue(str(row.get("queuename") or ""), qmap)
        row["total_calls"] = total
        row["sla"] = round(answered * 100 / total, 2) if total else 0
        row["abandon_rate"] = round(abandoned * 100 / total, 2) if total else 0
        row["timeout_rate"] = round(timeout * 100 / total, 2) if total else 0
        row["avg_hold"] = round(hold_sec / answered, 2) if answered else 0
        row["avg_talk"] = round(talk_sec / answered, 2) if answered else 0
        row["aht"] = round((hold_sec + talk_sec) / answered, 2) if answered else 0

    for row in top_agents:
        agent = str(row.get("agent") or "")
        calls = int(row.get("answered") or 0)
        hold = int(row.get("hold_sec") or 0)
        talk = int(row.get("talk_sec") or 0)
        row["agent_display"] = _display_agent(agent, amap)
        row["avg_hold"] = round(hold / calls, 2) if calls else 0
        row["avg_talk"] = round(talk / calls, 2) if calls else 0
        row["aht"] = round((hold + talk) / calls, 2) if calls else 0
        row["talk_min"] = round(talk / 60, 2)
        row["share_answered"] = round(calls * 100 / total_answered, 2) if total_answered else 0

    agent_filter_set = set()
    for raw in agents:
        for key in _agent_aliases(raw):
            agent_filter_set.add(key)

    operator_stats_map: Dict[str, Dict[str, Any]] = {}
    daily_direction_talk: Dict[str, Dict[str, int]] = {}
    total_incoming_talk_sec = 0
    total_outgoing_talk_sec = 0
    total_incoming_calls = 0
    total_outgoing_calls = 0

    for row in operator_cdr_rows:
        billsec = int(row.get("billsec") or 0)
        if billsec <= 0:
            continue

        queue_agent = str(row.get("queue_agent") or "").strip()
        operator_raw = queue_agent or str(row.get("dstchannel") or "").strip() or str(row.get("channel") or "").strip()
        if not operator_raw:
            src = str(row.get("src") or "").strip()
            dst = str(row.get("dst") or "").strip()
            if _is_internal_party(src):
                operator_raw = src
            elif _is_internal_party(dst):
                operator_raw = dst
        if not operator_raw:
            continue

        aliases = _agent_aliases(operator_raw)
        if not aliases:
            continue
        operator_key = aliases[-1]

        if agent_filter_set and not (set(aliases) & agent_filter_set):
            continue

        direction = _classify_call_direction(
            str(row.get("src") or ""),
            str(row.get("dst") or ""),
            str(row.get("dcontext") or ""),
            queue_agent,
        )
        day_label = str(row.get("calldate") or "")[:10]

        operator_stats_map.setdefault(
            operator_key,
            {
                "agent": operator_key,
                "agent_display": _display_agent(operator_raw, amap),
                "handled_calls": 0,
                "talk_sec_total": 0,
                "incoming_calls": 0,
                "incoming_talk_sec": 0,
                "outgoing_calls": 0,
                "outgoing_talk_sec": 0,
                "unknown_calls": 0,
                "unknown_talk_sec": 0,
            },
        )
        stat = operator_stats_map[operator_key]
        stat["handled_calls"] += 1
        stat["talk_sec_total"] += billsec
        if direction == "incoming":
            stat["incoming_calls"] += 1
            stat["incoming_talk_sec"] += billsec
            total_incoming_calls += 1
            total_incoming_talk_sec += billsec
        elif direction == "outgoing":
            stat["outgoing_calls"] += 1
            stat["outgoing_talk_sec"] += billsec
            total_outgoing_calls += 1
            total_outgoing_talk_sec += billsec
        else:
            stat["unknown_calls"] += 1
            stat["unknown_talk_sec"] += billsec

        if day_label:
            daily_direction_talk.setdefault(day_label, {"day": day_label, "incoming_talk_sec": 0, "outgoing_talk_sec": 0})
            if direction == "incoming":
                daily_direction_talk[day_label]["incoming_talk_sec"] += billsec
            elif direction == "outgoing":
                daily_direction_talk[day_label]["outgoing_talk_sec"] += billsec

    operator_duration_rows = sorted(
        operator_stats_map.values(),
        key=lambda item: (int(item.get("handled_calls") or 0), int(item.get("talk_sec_total") or 0)),
        reverse=True,
    )
    rank_calls = sorted(
        operator_duration_rows,
        key=lambda item: (int(item.get("handled_calls") or 0), int(item.get("talk_sec_total") or 0)),
        reverse=True,
    )
    rank_talk = sorted(
        operator_duration_rows,
        key=lambda item: (int(item.get("talk_sec_total") or 0), int(item.get("handled_calls") or 0)),
        reverse=True,
    )
    rank_by_calls_map = {str(row.get("agent")): idx for idx, row in enumerate(rank_calls, start=1)}
    rank_by_talk_map = {str(row.get("agent")): idx for idx, row in enumerate(rank_talk, start=1)}
    for row in operator_duration_rows:
        handled_calls = int(row.get("handled_calls") or 0)
        talk_sec_total = int(row.get("talk_sec_total") or 0)
        row["avg_talk_sec"] = round(talk_sec_total / handled_calls, 2) if handled_calls else 0
        row["talk_min_total"] = round(talk_sec_total / 60, 2)
        row["incoming_talk_min"] = round(int(row.get("incoming_talk_sec") or 0) / 60, 2)
        row["outgoing_talk_min"] = round(int(row.get("outgoing_talk_sec") or 0) / 60, 2)
        row["rank_by_calls"] = rank_by_calls_map.get(str(row.get("agent")), 0)
        row["rank_by_talk"] = rank_by_talk_map.get(str(row.get("agent")), 0)

    top_operator_by_calls = rank_calls[:20]
    top_operator_by_talk = rank_talk[:20]
    direction_daily_rows = [daily_direction_talk[key] for key in sorted(daily_direction_talk.keys())]
    for row in direction_daily_rows:
        row["incoming_talk_min"] = round(int(row.get("incoming_talk_sec") or 0) / 60, 2)
        row["outgoing_talk_min"] = round(int(row.get("outgoing_talk_sec") or 0) / 60, 2)

    daily_chart_input: List[Dict[str, Any]] = []
    for row in daily:
        daily_chart_input.append(
            {
                "day": row.get("day"),
                "total": int(row.get("answered") or 0) + int(row.get("unanswered") or 0),
            }
        )
    hourly_chart_input: List[Dict[str, Any]] = []
    for row in hourly:
        hourly_chart_input.append(
            {
                "hour": row.get("hour"),
                "total": int(row.get("answered") or 0) + int(row.get("unanswered") or 0),
            }
        )

    for row in top_callers:
        row["caller"] = str(row.get("caller") or "").strip()
        row["calls"] = int(row.get("calls") or 0)

    return {
        "start": start,
        "end": end,
        "kpi_total": total_calls,
        "kpi_answered": total_answered,
        "kpi_abandoned": total_abandoned,
        "kpi_timeout": total_timeout,
        "kpi_unanswered": total_unanswered,
        "kpi_sla": round(total_answered * 100 / total_calls, 2) if total_calls else 0,
        "kpi_abandon_rate": round(total_abandoned * 100 / total_calls, 2) if total_calls else 0,
        "kpi_timeout_rate": round(total_timeout * 100 / total_calls, 2) if total_calls else 0,
        "kpi_avg_hold": round(sum_hold_answered / total_answered, 2) if total_answered else 0,
        "kpi_avg_talk": round(sum_talk_answered / total_answered, 2) if total_answered else 0,
        "kpi_aht": round((sum_hold_answered + sum_talk_answered) / total_answered, 2) if total_answered else 0,
        "kpi_total_talk_min": round(sum_talk_answered / 60, 2),
        "kpi_active_queues": len([x for x in per_queue if int(x.get("total_calls") or 0) > 0]),
        "kpi_active_agents": len([x for x in top_agents if int(x.get("answered") or 0) > 0]),
        "kpi_incoming_talk_min": round(total_incoming_talk_sec / 60, 2),
        "kpi_outgoing_talk_min": round(total_outgoing_talk_sec / 60, 2),
        "kpi_incoming_calls": total_incoming_calls,
        "kpi_outgoing_calls": total_outgoing_calls,
        "daily": daily,
        "hourly": hourly,
        "per_queue": per_queue,
        "top_agents": top_agents,
        "top_callers": top_callers,
        "operator_duration_rows": operator_duration_rows,
        "operator_rank_by_calls": top_operator_by_calls,
        "operator_rank_by_talk": top_operator_by_talk,
        "daily_direction_talk": direction_daily_rows,
        "daily_line_chart": _line_chart(daily_chart_input, "day", "total", max_items=31),
        "hourly_bar_chart": _bar_chart(hourly_chart_input, "hour", "total", max_items=24),
        "queue_calls_chart": _bar_chart(per_queue, "queue_display", "total_calls", max_items=16),
        "agent_answered_chart": _bar_chart(top_agents, "agent_display", "answered", max_items=16),
        "frequent_callers_chart": _bar_chart(top_callers, "caller", "calls", max_items=16),
        "operator_total_calls_chart": _bar_chart(top_operator_by_calls, "agent_display", "handled_calls", max_items=16),
        "operator_total_talk_chart": _bar_chart(top_operator_by_talk, "agent_display", "talk_min_total", max_items=16),
        "incoming_talk_daily_chart": _line_chart(direction_daily_rows, "day", "incoming_talk_min", max_items=31),
        "outgoing_talk_daily_chart": _line_chart(direction_daily_rows, "day", "outgoing_talk_min", max_items=31),
    }


def _payout_dataset(request: HttpRequest) -> Dict[str, Any]:
    start, end = _interval_from_request(request)
    queues = _get_param_list(request, "queues")
    agents = _get_param_list(request, "agents")
    qmap = _queue_map()
    amap = _agent_map()
    rate_map = _payout_rate_map()
    general = _get_general_settings()
    default_rate = Decimal(general.default_payout_rate_per_minute)
    currency_code = general.currency_code

    rows = _fetch_queuelog_rows(start, end, queues, agents, ["COMPLETECALLER", "COMPLETEAGENT"])

    by_agent: Dict[str, Dict[str, Any]] = {}
    by_queue: Dict[str, Dict[str, Any]] = {}
    total_talk_sec = 0
    total_payout = Decimal("0.00")

    for row in rows:
        agent_raw = str(row.get("agent") or "")
        queue_raw = str(row.get("queuename") or "")
        talk_sec = int(row.get("data2") or 0)
        if not agent_raw or talk_sec <= 0:
            continue

        aliases = _agent_aliases(agent_raw)
        agent_key = aliases[-1] if aliases else agent_raw
        rate = default_rate
        for key in aliases:
            if key in rate_map:
                rate = rate_map[key]
                break
        payout = ((Decimal(talk_sec) / Decimal("60")) * rate).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        by_agent.setdefault(
            agent_key,
            {
                "agent": agent_key,
                "agent_display": _display_agent(agent_raw, amap),
                "calls": 0,
                "talk_sec": 0,
                "talk_min": Decimal("0.00"),
                "rate_per_minute": rate,
                "payout_total": Decimal("0.00"),
            },
        )
        by_agent[agent_key]["calls"] += 1
        by_agent[agent_key]["talk_sec"] += talk_sec
        by_agent[agent_key]["talk_min"] = (Decimal(by_agent[agent_key]["talk_sec"]) / Decimal("60")).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
        by_agent[agent_key]["rate_per_minute"] = rate
        by_agent[agent_key]["payout_total"] = (by_agent[agent_key]["payout_total"] + payout).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )

        queue_key = f"{queue_raw}::{agent_key}"
        by_queue.setdefault(
            queue_key,
            {
                "queue": queue_raw,
                "queue_display": _display_queue(queue_raw, qmap),
                "agent": agent_key,
                "agent_display": _display_agent(agent_raw, amap),
                "calls": 0,
                "talk_sec": 0,
                "talk_min": Decimal("0.00"),
                "rate_per_minute": rate,
                "payout_total": Decimal("0.00"),
            },
        )
        by_queue[queue_key]["calls"] += 1
        by_queue[queue_key]["talk_sec"] += talk_sec
        by_queue[queue_key]["talk_min"] = (Decimal(by_queue[queue_key]["talk_sec"]) / Decimal("60")).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
        by_queue[queue_key]["rate_per_minute"] = rate
        by_queue[queue_key]["payout_total"] = (by_queue[queue_key]["payout_total"] + payout).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )

        total_talk_sec += talk_sec
        total_payout += payout

    by_agent_rows = sorted(by_agent.values(), key=lambda x: x["payout_total"], reverse=True)
    by_queue_rows = sorted(by_queue.values(), key=lambda x: x["payout_total"], reverse=True)
    with_rate = len([x for x in by_agent_rows if x["rate_per_minute"] > 0])

    return {
        "start": start,
        "end": end,
        "rows_agent": by_agent_rows,
        "rows_queue": by_queue_rows,
        "total_calls": sum(int(x["calls"]) for x in by_agent_rows),
        "total_talk_min": (Decimal(total_talk_sec) / Decimal("60")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP),
        "total_payout": total_payout.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP),
        "agents_with_rate": with_rate,
        "agents_without_rate": max(0, len(by_agent_rows) - with_rate),
        "currency_code": currency_code,
        "default_payout_rate_per_minute": default_rate.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP),
    }


def _add_progress(rows: List[Dict[str, Any]], field: str, target: str = "bar_pct") -> List[Dict[str, Any]]:
    max_value = max([float(r.get(field) or 0) for r in rows], default=0.0)
    if max_value <= 0:
        for row in rows:
            row[target] = 0
        return rows
    for row in rows:
        row[target] = round((float(row.get(field) or 0) * 100.0) / max_value, 2)
    return rows


def _bar_chart(items: List[Dict[str, Any]], label_key: str, value_key: str, max_items: int = 16) -> Dict[str, Any]:
    src = items[:max_items]
    values = [float(x.get(value_key) or 0) for x in src]
    max_value = max(values, default=0.0) or 1.0
    count = max(1, len(src))
    width = 640
    height = 240
    pad = 28
    inner_w = width - (pad * 2)
    inner_h = height - (pad * 2)
    slot = inner_w / count
    bar_w = max(8, slot * 0.72)
    bars: List[Dict[str, Any]] = []
    for idx, item in enumerate(src):
        value = float(item.get(value_key) or 0)
        ratio = value / max_value
        h = ratio * inner_h
        x = pad + (idx * slot) + ((slot - bar_w) / 2)
        y = pad + inner_h - h
        bars.append(
            {
                "x": round(x, 2),
                "y": round(y, 2),
                "w": round(bar_w, 2),
                "h": round(h, 2),
                "label": str(item.get(label_key) or ""),
                "value": item.get(value_key) or 0,
            }
        )
    return {"w": width, "h": height, "bars": bars, "max_value": round(max_value, 2)}


def _line_chart(items: List[Dict[str, Any]], label_key: str, value_key: str, max_items: int = 30) -> Dict[str, Any]:
    src = items[:max_items]
    values = [float(x.get(value_key) or 0) for x in src]
    max_value = max(values, default=0.0) or 1.0
    count = max(2, len(src))
    width = 640
    height = 240
    pad = 28
    inner_w = width - (pad * 2)
    inner_h = height - (pad * 2)
    step = inner_w / max(1, count - 1)
    points: List[Dict[str, Any]] = []
    poly_parts: List[str] = []
    for idx, item in enumerate(src):
        value = float(item.get(value_key) or 0)
        ratio = value / max_value
        x = pad + (idx * step)
        y = pad + inner_h - (ratio * inner_h)
        points.append(
            {
                "x": round(x, 2),
                "y": round(y, 2),
                "label": str(item.get(label_key) or ""),
                "value": item.get(value_key) or 0,
            }
        )
        poly_parts.append(f"{round(x,2)},{round(y,2)}")
    return {
        "w": width,
        "h": height,
        "points": points,
        "polyline": " ".join(poly_parts),
        "max_value": round(max_value, 2),
    }


def _dashboard_traffic_dataset(request: HttpRequest) -> Dict[str, Any]:
    data = _analytics_dataset(request)
    daily = data.get("daily", [])
    hourly = data.get("hourly", [])
    for row in daily:
        row["total"] = int(row.get("answered") or 0) + int(row.get("unanswered") or 0)
    for row in hourly:
        row["total"] = int(row.get("answered") or 0) + int(row.get("unanswered") or 0)
    data["daily_line_chart"] = _line_chart(daily, "day", "total", max_items=31)
    data["hourly_bar_chart"] = _bar_chart(hourly, "hour", "total", max_items=24)
    data["daily"] = daily
    data["hourly"] = hourly
    return data


def _dashboard_queues_dataset(request: HttpRequest) -> Dict[str, Any]:
    data = _analytics_dataset(request)
    rows = data.get("per_queue", [])
    data["queues_calls_chart"] = _bar_chart(rows, "queue_display", "total_calls")
    data["queues_sla_chart"] = _bar_chart(rows, "queue_display", "sla")
    data["per_queue"] = rows
    return data


def _dashboard_operators_dataset(request: HttpRequest) -> Dict[str, Any]:
    analytics = _analytics_dataset(request)
    payouts = _payout_dataset(request)
    payout_map: Dict[str, Dict[str, Any]] = {str(r.get("agent")): r for r in payouts.get("rows_agent", [])}

    rows: List[Dict[str, Any]] = []
    for row in analytics.get("top_agents", []):
        aliases = _agent_aliases(str(row.get("agent") or ""))
        agent_key = aliases[-1] if aliases else str(row.get("agent") or "")
        pay = payout_map.get(agent_key, {})
        rows.append(
            {
                "agent": agent_key,
                "agent_display": row.get("agent_display"),
                "answered": int(row.get("answered") or 0),
                "talk_min": row.get("talk_min", 0),
                "avg_talk": row.get("avg_talk", 0),
                "aht": row.get("aht", 0),
                "share_answered": row.get("share_answered", 0),
                "rate_per_minute": pay.get("rate_per_minute", Decimal("0.00")),
                "payout_total": pay.get("payout_total", Decimal("0.00")),
            }
        )
    analytics["operators_talk_chart"] = _bar_chart(rows, "agent_display", "talk_min")
    analytics["operators_payout_chart"] = _bar_chart(rows, "agent_display", "payout_total")
    analytics["rows_operators"] = rows
    analytics["total_payout"] = payouts.get("total_payout", Decimal("0.00"))
    analytics["currency_code"] = payouts.get("currency_code", "UZS")
    return analytics


def _build_ami_snapshot(request: HttpRequest | None = None, filters: Dict[str, str] | None = None) -> Dict[str, Any]:
    def _snapshot_filter_value(key: str) -> str:
        if request is not None:
            return _filter_value(request, key)
        if filters is not None:
            return filters.get(key, "")
        return ""

    qmap = _queue_map()
    amap = _agent_map()
    if request is not None:
        queues_filter = set(_get_param_list(request, "queues"))
    else:
        queues_filter = set(_normalize_list(_snapshot_filter_value("queues")))
    channel_filter = (_snapshot_filter_value("channel") or "").strip().lower()
    caller_filter = (_snapshot_filter_value("caller") or "").strip().lower()

    settings = GeneralSettings.objects.first()
    if not settings or not settings.ami_host:
        return {
            "queue_summary": [],
            "active_calls": [],
            "active_calls_count": 0,
            "waiting_calls_count": 0,
            "active_operators_count": 0,
            "ami_error": i18n_tr("AMI не настроен"),
        }

    manager = AMIManager(
        host=settings.ami_host,
        port=settings.ami_port,
        username=settings.ami_user,
        secret=settings.ami_password,
    )

    if not manager.connect():
        return {
            "queue_summary": [],
            "active_calls": [],
            "active_calls_count": 0,
            "waiting_calls_count": 0,
            "active_operators_count": 0,
            "ami_error": i18n_tr("Нет соединения с AMI"),
        }

    try:
        summary_raw = manager.queue_summary().get("summary", [])
        channels_raw = manager.core_show_channels().get("channels", [])
    finally:
        manager.disconnect()

    queue_summary: List[Dict[str, Any]] = []
    for row in summary_raw:
        if str(row.get("Event", "")) != "QueueSummary":
            continue
        queue_system = str(row.get("Queue", ""))
        if queues_filter and queue_system not in queues_filter:
            continue
        queue_summary.append(
            {
                "queue": queue_system,
                "queue_display": _display_queue(queue_system, qmap),
                "logged_in": row.get("LoggedIn", "0"),
                "available": row.get("Available", "0"),
                "callers": row.get("Callers", "0"),
                "hold_time": row.get("HoldTime", "0"),
                "longest_hold": row.get("LongestHoldTime", "0"),
            }
        )

    deduped_channels: Dict[str, Dict[str, Any]] = {}
    fallback_idx = 0
    for row in channels_raw:
        if str(row.get("Event", "")) != "CoreShowChannel":
            continue
        channel = str(row.get("Channel", ""))
        if channel.startswith("Message/"):
            continue
        linkedid = str(row.get("Linkedid") or row.get("LinkedId") or "").strip()
        bridge_id = str(row.get("BridgeId") or row.get("BridgeID") or row.get("BridgeUniqueid") or "").strip()
        dedupe_key = linkedid or bridge_id
        if not dedupe_key:
            fallback_idx += 1
            dedupe_key = f"row:{fallback_idx}:{channel}"

        current = deduped_channels.get(dedupe_key)
        if current is None or _channel_row_rank(row) > _channel_row_rank(current):
            deduped_channels[dedupe_key] = row

    active_calls: List[Dict[str, Any]] = []
    active_operator_ids = set()
    for row in deduped_channels.values():
        channel = str(row.get("Channel", ""))
        caller = str(row.get("CallerIDNum", ""))
        connected = str(row.get("ConnectedLineNum", ""))
        callid = str(row.get("Linkedid") or row.get("LinkedId") or row.get("BridgeId") or row.get("BridgeID") or "").strip()
        if channel_filter and channel_filter not in channel.lower():
            continue
        caller_human = _human_party(caller, amap)
        connected_human = _human_party(connected, amap)
        haystack = " ".join([caller, connected, caller_human, connected_human]).lower()
        if caller_filter and caller_filter not in haystack:
            continue
        active_calls.append(
            {
                "callid": callid,
                "channel": _human_channel(channel, amap),
                "caller": caller_human,
                "connected": connected_human,
                "duration": row.get("Duration", ""),
                "application": row.get("Application", ""),
            }
        )
        for candidate in (channel, caller, connected):
            ext = _extract_operator_ext(candidate)
            if ext:
                active_operator_ids.add(ext)

    waiting_calls_total = 0
    for row in queue_summary:
        try:
            waiting_calls_total += int(row.get("callers") or 0)
        except (TypeError, ValueError):
            continue

    return {
        "queue_summary": queue_summary,
        "active_calls": active_calls,
        "active_calls_count": len(active_calls),
        "waiting_calls_count": waiting_calls_total,
        "active_operators_count": len(active_operator_ids),
        "ami_error": "",
    }


def _draw_table_pdf(title: str, headers: List[str], rows: List[List[Any]]) -> bytes:
    output = BytesIO()
    pdf = canvas.Canvas(output, pagesize=landscape(A4))
    page_width, page_height = landscape(A4)

    pdf.setFillColor(colors.HexColor("#111827"))
    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(12 * mm, page_height - 12 * mm, title)

    y = page_height - 24 * mm
    x = 10 * mm
    col_width = (page_width - 20 * mm) / max(1, len(headers))

    pdf.setFillColor(colors.HexColor("#1F2937"))
    pdf.rect(x, y, page_width - 20 * mm, 8 * mm, fill=1, stroke=0)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 8)
    for i, header in enumerate(headers):
        pdf.drawString(x + (i * col_width) + 1.5 * mm, y + 2.5 * mm, str(header)[:30])

    y -= 7 * mm
    pdf.setFont("Helvetica", 7)
    pdf.setFillColor(colors.black)
    for row in rows[:180]:
        if y < 12 * mm:
            pdf.showPage()
            y = page_height - 14 * mm
            pdf.setFont("Helvetica", 7)
        for i, value in enumerate(row):
            pdf.drawString(x + (i * col_width) + 1.5 * mm, y, str(value)[:30])
        y -= 5 * mm

    pdf.save()
    output.seek(0)
    return output.read()


def _draw_bar_plot_on_canvas(pdf, title: str, x: float, y: float, w: float, h: float, labels: List[str], values: List[float]) -> None:
    pdf.setStrokeColor(colors.HexColor("#334155"))
    pdf.setFillColor(colors.HexColor("#0B1220"))
    pdf.rect(x, y, w, h, fill=1, stroke=1)
    pdf.setFillColor(colors.HexColor("#CBD5E1"))
    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawString(x + 4, y + h - 11, title[:80])

    if not values:
        pdf.setFont("Helvetica", 8)
        pdf.drawString(x + 4, y + 6, "No data")
        return

    pad = 8
    inner_x = x + pad
    inner_y = y + 18
    inner_w = w - (pad * 2)
    inner_h = h - 28
    max_val = max(values) if max(values) > 0 else 1
    count = max(1, len(values))
    slot = inner_w / count
    bar_w = max(3, slot * 0.68)

    pdf.setFillColor(colors.HexColor("#60A5FA"))
    for idx, value in enumerate(values):
        bar_h = (value / max_val) * inner_h
        bx = inner_x + (idx * slot) + ((slot - bar_w) / 2)
        by = inner_y
        pdf.rect(bx, by, bar_w, bar_h, fill=1, stroke=0)

    pdf.setFillColor(colors.HexColor("#94A3B8"))
    pdf.setFont("Helvetica", 6)
    max_labels = min(len(labels), 10)
    step = max(1, len(labels) // max_labels)
    for idx in range(0, len(labels), step):
        lx = inner_x + (idx * slot)
        pdf.drawString(lx, y + 4, str(labels[idx])[:8])


def _draw_line_plot_on_canvas(pdf, title: str, x: float, y: float, w: float, h: float, labels: List[str], values: List[float]) -> None:
    pdf.setStrokeColor(colors.HexColor("#334155"))
    pdf.setFillColor(colors.HexColor("#0B1220"))
    pdf.rect(x, y, w, h, fill=1, stroke=1)
    pdf.setFillColor(colors.HexColor("#CBD5E1"))
    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawString(x + 4, y + h - 11, title[:80])

    if not values:
        pdf.setFont("Helvetica", 8)
        pdf.drawString(x + 4, y + 6, "No data")
        return

    pad = 10
    inner_x = x + pad
    inner_y = y + 18
    inner_w = w - (pad * 2)
    inner_h = h - 28
    max_val = max(values) if max(values) > 0 else 1
    count = max(2, len(values))
    step_x = inner_w / max(1, count - 1)
    points = []
    for idx, value in enumerate(values):
        px = inner_x + (idx * step_x)
        py = inner_y + ((value / max_val) * inner_h)
        points.append((px, py))

    pdf.setStrokeColor(colors.HexColor("#22D3EE"))
    pdf.setLineWidth(1.5)
    for idx in range(len(points) - 1):
        x1, y1 = points[idx]
        x2, y2 = points[idx + 1]
        pdf.line(x1, y1, x2, y2)
    pdf.setFillColor(colors.HexColor("#38BDF8"))
    for px, py in points:
        pdf.circle(px, py, 1.7, fill=1, stroke=0)

    pdf.setFillColor(colors.HexColor("#94A3B8"))
    pdf.setFont("Helvetica", 6)
    max_labels = min(len(labels), 10)
    step = max(1, len(labels) // max_labels)
    for idx in range(0, len(labels), step):
        lx = inner_x + (idx * step_x)
        pdf.drawString(lx, y + 4, str(labels[idx])[:8])


def _draw_plots_pdf(
    title: str,
    plots: List[Dict[str, Any]],
    tables: List[Dict[str, Any]] | None = None,
) -> bytes:
    output = BytesIO()
    pdf = canvas.Canvas(output, pagesize=landscape(A4))
    page_w, page_h = landscape(A4)
    margin = 10 * mm
    gap = 6 * mm
    plot_w = (page_w - (margin * 2) - gap) / 2
    plot_h = (page_h - (margin * 2) - 18 * mm - gap) / 2
    base_y = margin
    top_y = base_y + plot_h + gap

    positions = [
        (margin, top_y),
        (margin + plot_w + gap, top_y),
        (margin, base_y),
        (margin + plot_w + gap, base_y),
    ]

    for page_index in range(0, len(plots), 4):
        if page_index > 0:
            pdf.showPage()
        pdf.setFillColor(colors.HexColor("#111827"))
        pdf.setFont("Helvetica-Bold", 14)
        page_no = (page_index // 4) + 1
        pdf.drawString(12 * mm, page_h - 12 * mm, f"{title[:96]} (charts {page_no})")

        for idx, plot in enumerate(plots[page_index : page_index + 4]):
            x, y = positions[idx]
            labels = [str(val) for val in plot.get("labels", [])]
            values = [float(val or 0) for val in plot.get("values", [])]
            if plot.get("type") == "line":
                _draw_line_plot_on_canvas(pdf, plot.get("title", "Plot"), x, y, plot_w, plot_h, labels, values)
            else:
                _draw_bar_plot_on_canvas(pdf, plot.get("title", "Plot"), x, y, plot_w, plot_h, labels, values)

    for table in (tables or []):
        headers = [str(h) for h in table.get("headers", [])]
        rows = table.get("rows", []) or []
        if not headers:
            continue

        col_count = max(1, len(headers))
        col_width = (page_w - 20 * mm) / col_count
        x = 10 * mm
        y = page_h - 24 * mm
        rows_printed = 0

        while rows_printed < len(rows) or rows_printed == 0:
            pdf.showPage()
            pdf.setFillColor(colors.HexColor("#111827"))
            pdf.setFont("Helvetica-Bold", 14)
            pdf.drawString(12 * mm, page_h - 12 * mm, str(table.get("title") or "Table")[:120])

            header_y = page_h - 24 * mm
            pdf.setFillColor(colors.HexColor("#1F2937"))
            pdf.rect(x, header_y, page_w - 20 * mm, 8 * mm, fill=1, stroke=0)
            pdf.setFillColor(colors.white)
            pdf.setFont("Helvetica-Bold", 8)
            for i, header in enumerate(headers):
                pdf.drawString(x + (i * col_width) + 1.5 * mm, header_y + 2.5 * mm, header[:26])

            y = header_y - 7 * mm
            pdf.setFont("Helvetica", 7)
            pdf.setFillColor(colors.black)
            if not rows:
                pdf.drawString(x + 1.5 * mm, y, "No data")
                break

            while rows_printed < len(rows):
                if y < 12 * mm:
                    break
                row = rows[rows_printed]
                for i, value in enumerate(row[:col_count]):
                    pdf.drawString(x + (i * col_width) + 1.5 * mm, y, str(value)[:26])
                rows_printed += 1
                y -= 5 * mm

    pdf.save()
    output.seek(0)
    return output.read()


def _recording_file_by_uniqueid(uniqueid: str) -> str:
    with connections["default"].cursor() as cursor:
        cursor.execute("SELECT recordingfile FROM cdr WHERE uniqueid = %s", [uniqueid])
        row = cursor.fetchone()
    if not row or not row[0]:
        raise Http404("Recording not found")
    return str(row[0])


def _resolve_recording_local_path(recording_path: str) -> Path | None:
    candidates: List[Path] = []
    raw = Path(recording_path)
    if raw.is_absolute():
        candidates.append(raw)
    monitor_base = Path(getattr(django_settings, "ASTERISK_MONITOR_PATH", "/var/spool/asterisk/monitor/"))
    candidates.append(monitor_base / recording_path)
    candidates.append(monitor_base / raw.name)

    for path in candidates:
        try:
            resolved = path.resolve()
            if resolved.is_file():
                return resolved
        except Exception:
            continue
    return None


def _parse_range_header(range_header: str, file_size: int) -> tuple[int, int] | None:
    if not range_header or not range_header.startswith("bytes="):
        return None
    value = range_header.replace("bytes=", "", 1).strip()
    if "," in value:
        value = value.split(",", 1)[0].strip()
    if "-" not in value:
        return None
    start_raw, end_raw = value.split("-", 1)
    try:
        if start_raw == "":
            length = int(end_raw)
            if length <= 0:
                return None
            start = max(0, file_size - length)
            end = file_size - 1
            return start, end
        start = int(start_raw)
        end = int(end_raw) if end_raw else file_size - 1
        if start > end or start < 0 or end >= file_size:
            return None
        return start, end
    except ValueError:
        return None


def _stream_file_range(path: Path, start: int, end: int, chunk_size: int = 8192):
    with path.open("rb") as handle:
        handle.seek(start)
        remaining = end - start + 1
        while remaining > 0:
            chunk = handle.read(min(chunk_size, remaining))
            if not chunk:
                break
            remaining -= len(chunk)
            yield chunk


def web_login(request: HttpRequest) -> HttpResponse:
    if request.user.is_authenticated:
        return redirect("home")

    context: Dict[str, Any] = {}
    if request.method == "POST":
        username = (request.POST.get("username") or "").strip()
        password = request.POST.get("password") or ""
        user = authenticate(request, username=username, password=password)
        if user is None:
            context["error"] = i18n_tr("Неверный логин или пароль")
        else:
            login(request, user)
            return redirect("home")

    return render(request, "stats/login.html", context)


@login_required
def web_logout(request: HttpRequest) -> HttpResponse:
    logout(request)
    return redirect("web-login")


@login_required
def home(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)
    return render(request, "stats/home.html", _base_context(request))


@login_required
def report_summary_page(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)
    context = _base_context(request)
    context.update(_summary_dataset(request))
    return render(request, "stats/reports/summary_page.html", context)


@login_required
def report_answered_page(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)
    context = _base_context(request)
    context.update(_answered_dataset(request))
    return render(request, "stats/reports/answered_page.html", context)


@login_required
def report_unanswered_page(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)
    context = _base_context(request)
    context.update(_unanswered_dataset(request))
    return render(request, "stats/reports/unanswered_page.html", context)


@login_required
def report_cdr_page(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)
    context = _base_context(request)
    context.update(_cdr_dataset(request))
    return render(request, "stats/reports/cdr_page.html", context)


@login_required
def call_detail_page(request: HttpRequest, callid: str) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)
    if not callid or not all(c.isalnum() or c in ".-_:" for c in callid):
        raise Http404("Invalid callid")

    notice = ""
    error = ""
    if request.method == "POST":
        action = (request.POST.get("action") or "").strip().lower()
        if action == "transcribe":
            ok, message = _transcribe_call(callid)
            if ok:
                notice = message
            else:
                error = message

    dataset = _call_detail_dataset(callid)
    should_autotranscribe = (
        request.method != "POST"
        and dataset.get("transcription_configured")
        and bool(dataset.get("cdr") and dataset["cdr"].get("has_recording"))
        and dataset.get("transcription") is None
    )
    if should_autotranscribe:
        ok, message = _transcribe_call(callid)
        if ok:
            notice = message
        else:
            error = message
        dataset = _call_detail_dataset(callid)

    context = _base_context(request)
    context.update(dataset)
    context["notice"] = notice
    context["error"] = error
    return render(request, "stats/reports/call_detail_page.html", context)


@login_required
def realtime_page(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)
    context = _base_context(request)
    context.update(_build_ami_snapshot(request))
    return render(request, "stats/realtime_page.html", context)


@login_required
def analytics_page(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)
    context = _base_context(request)
    context.update(_analytics_dataset(request))
    return render(request, "stats/analytics_page.html", context)


@login_required
def payouts_page(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)
    context = _base_context(request)
    context.update(_payout_dataset(request))
    return render(request, "stats/payouts_page.html", context)


@login_required
def dashboards_page(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)
    return render(request, "stats/dashboards/index_page.html", _base_context(request))


@login_required
def dashboard_traffic_page(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)
    context = _base_context(request)
    context.update(_dashboard_traffic_dataset(request))
    return render(request, "stats/dashboards/traffic_page.html", context)


@login_required
def dashboard_queues_page(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)
    context = _base_context(request)
    context.update(_dashboard_queues_dataset(request))
    return render(request, "stats/dashboards/queues_page.html", context)


@login_required
def dashboard_operators_page(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)
    context = _base_context(request)
    context.update(_dashboard_operators_dataset(request))
    return render(request, "stats/dashboards/operators_page.html", context)


@login_required
def settings_page(request: HttpRequest) -> HttpResponse:
    if not _admin_allowed(request):
        return HttpResponse("forbidden", status=403)

    settings_obj = _get_general_settings()
    notice = ""
    if request.method == "POST":
        action = request.POST.get("action", "save")
        section = request.POST.get("section", "")
        mapping_type = request.POST.get("mapping_type", "")
        system_name = (request.POST.get("system_name") or "").strip()
        display_name = (request.POST.get("display_name") or "").strip()

        if section == "general":
            settings_obj.currency_code = (request.POST.get("currency_code") or settings_obj.currency_code or "UZS").strip().upper()
            settings_obj.ui_language = (request.POST.get("ui_language") or settings_obj.ui_language or "ru").strip().lower()
            settings_obj.default_payout_rate_per_minute = request.POST.get("default_payout_rate_per_minute") or settings_obj.default_payout_rate_per_minute
            settings_obj.sla_target_percent = request.POST.get("sla_target_percent") or settings_obj.sla_target_percent
            settings_obj.sla_target_wait_seconds = request.POST.get("sla_target_wait_seconds") or settings_obj.sla_target_wait_seconds
            settings_obj.transcription_url = (request.POST.get("transcription_url") or settings_obj.transcription_url or "").strip()
            settings_obj.transcription_api_key = (request.POST.get("transcription_api_key") or settings_obj.transcription_api_key or "").strip()

            start_raw = (request.POST.get("business_day_start") or "").strip()
            end_raw = (request.POST.get("business_day_end") or "").strip()
            if start_raw:
                try:
                    settings_obj.business_day_start = datetime.strptime(start_raw, "%H:%M").time()
                except ValueError:
                    pass
            if end_raw:
                try:
                    settings_obj.business_day_end = datetime.strptime(end_raw, "%H:%M").time()
                except ValueError:
                    pass

            settings_obj.save()
            notice = i18n_tr("Общие настройки сохранены")
        elif mapping_type == "queue" and system_name:
            if action == "delete":
                QueueDisplayMapping.objects.filter(queue_system_name=system_name).delete()
                notice = i18n_tr("Маппинг очереди удалён")
            elif display_name:
                QueueDisplayMapping.objects.update_or_create(
                    queue_system_name=system_name,
                    defaults={"queue_display_name": display_name},
                )
                notice = i18n_tr("Маппинг очереди сохранён")
        elif mapping_type == "agent" and system_name:
            if action == "delete":
                AgentDisplayMapping.objects.filter(agent_system_name=system_name).delete()
                notice = i18n_tr("Маппинг оператора удалён")
            elif display_name:
                AgentDisplayMapping.objects.update_or_create(
                    agent_system_name=system_name,
                    defaults={"agent_display_name": display_name},
                )
                notice = i18n_tr("Маппинг оператора сохранён")
        elif mapping_type == "rate" and system_name:
            if action == "delete":
                OperatorPayoutRate.objects.filter(agent_system_name=system_name).delete()
                notice = i18n_tr("Ставка оператора удалена")
            else:
                rate_value = request.POST.get("rate_per_minute", "0").strip()
                try:
                    rate = float(rate_value)
                except ValueError:
                    rate = 0.0
                OperatorPayoutRate.objects.update_or_create(
                    agent_system_name=system_name,
                    defaults={"rate_per_minute": rate},
                )
                notice = i18n_tr("Ставка оператора сохранена")

    query = (request.GET.get("q") or "").strip()
    queue_qs = QueueDisplayMapping.objects.all()
    agent_qs = AgentDisplayMapping.objects.all()
    payout_qs = OperatorPayoutRate.objects.all()
    if query:
        queue_qs = queue_qs.filter(queue_system_name__icontains=query) | queue_qs.filter(queue_display_name__icontains=query)
        agent_qs = agent_qs.filter(agent_system_name__icontains=query) | agent_qs.filter(agent_display_name__icontains=query)
        payout_qs = payout_qs.filter(agent_system_name__icontains=query)

    context = _base_context(request)
    payout_rows: List[Dict[str, Any]] = []
    amap = _agent_map()
    for row in payout_qs:
        agent_system_name = str(row.agent_system_name or "")
        payout_rows.append(
            {
                "agent_system_name": agent_system_name,
                "agent_display_name": _display_agent(agent_system_name, amap),
                "rate_per_minute": row.rate_per_minute,
            }
        )

    context.update(
        {
            "notice": notice,
            "settings_obj": settings_obj,
            "language_choices": [("uz", "uz"), ("ru", "ru"), ("en", "en")],
            "currency_choices": ["UZS", "USD", "RUB", "EUR"],
            "queue_mappings": queue_qs,
            "agent_mappings": agent_qs,
            "payout_rates": payout_rows,
        }
    )
    return render(request, "stats/settings_page.html", context)


@login_required
def mappings_page(request: HttpRequest) -> HttpResponse:
    return redirect("settings-page")


@login_required
def recording_stream(request: HttpRequest, uniqueid: str) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)

    if not uniqueid or not all(c.isalnum() or c in ".-" for c in uniqueid):
        raise Http404("Invalid uniqueid")

    recording_path = _recording_file_by_uniqueid(uniqueid)
    local_path = _resolve_recording_local_path(recording_path)
    if local_path:
        file_size = local_path.stat().st_size
        content_type, _ = mimetypes.guess_type(str(local_path))
        content_type = content_type or "audio/wav"
        range_header = request.headers.get("Range") or request.META.get("HTTP_RANGE", "")
        byte_range = _parse_range_header(range_header, file_size)
        if byte_range:
            start, end = byte_range
            response = StreamingHttpResponse(
                _stream_file_range(local_path, start, end),
                status=206,
                content_type=content_type,
            )
            response["Content-Range"] = f"bytes {start}-{end}/{file_size}"
            response["Content-Length"] = str(end - start + 1)
        else:
            response = FileResponse(local_path.open("rb"), content_type=content_type)
            response["Content-Length"] = str(file_size)
        response["Accept-Ranges"] = "bytes"
        response["Content-Disposition"] = f'inline; filename="{os.path.basename(str(local_path))}"'
        response["Cache-Control"] = "no-store"
        return response

    conf = GeneralSettings.objects.first()
    if conf and conf.download_url:
        params = {"url": recording_path, "token": conf.download_token}
        auth = (conf.download_user, conf.download_password)
        headers = {}
        if request.headers.get("Range"):
            headers["Range"] = request.headers["Range"]
        try:
            upstream = requests.get(
                conf.download_url,
                params=params,
                auth=auth,
                headers=headers,
                stream=True,
                timeout=(5, 30),
            )
            upstream.raise_for_status()
        except requests.RequestException as exc:
            raise Http404(f"Failed to fetch recording: {exc}")

        response = StreamingHttpResponse(
            upstream.iter_content(chunk_size=8192),
            content_type=upstream.headers.get("Content-Type", "audio/mpeg"),
            status=upstream.status_code,
            reason=upstream.reason,
        )
        for key in ("Content-Length", "Content-Range", "Accept-Ranges"):
            if upstream.headers.get(key):
                response[key] = upstream.headers[key]
        response["Content-Disposition"] = f'inline; filename="{uniqueid}.wav"'
        response["Cache-Control"] = "no-store"
        return response

    raise Http404("Recording not available")


@login_required
def export_answered_excel(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)

    data = _answered_dataset(request)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Answered"
    sheet.append(["time", "caller", "queue", "agent", "hold_sec", "talk_sec"])
    for row in data["rows"]:
        sheet.append([row["time"], row.get("caller", ""), row["queue_display"], row["agent_display"], row["hold"], row["talk"]])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="answered_report.xlsx"'
    return response


@login_required
def export_unanswered_excel(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)

    data = _unanswered_dataset(request)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Unanswered"
    sheet.append(["time", "caller", "queue", "event", "start_pos", "end_pos", "wait_sec"])
    for row in data["rows"]:
        sheet.append(
            [
                row["time"],
                row.get("caller", ""),
                row["queue_display"],
                row["event"],
                row["start_pos"],
                row["end_pos"],
                row["wait_sec"],
            ]
        )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="unanswered_report.xlsx"'
    return response


@login_required
def export_cdr_excel(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)

    data = _cdr_dataset(request)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "CDR"
    sheet.append(["uniqueid", "calldate", "operator", "src", "dst", "duration", "billsec", "disposition", "recordingfile"])
    for row in data["rows"]:
        sheet.append(
            [
                row.get("uniqueid"),
                row.get("calldate"),
                row.get("operator_display"),
                row.get("src"),
                row.get("dst_display"),
                row.get("duration"),
                row.get("billsec"),
                row.get("disposition"),
                row.get("recordingfile"),
            ]
        )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="cdr_report.xlsx"'
    return response


@login_required
def export_answered_pdf(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)

    data = _answered_dataset(request)
    pdf_data = _draw_table_pdf(
        "Answered Report",
        ["time", "caller", "queue", "agent", "hold_sec", "talk_sec"],
        [[r["time"], r.get("caller", ""), r["queue_display"], r["agent_display"], r["hold"], r["talk"]] for r in data["rows"]],
    )
    response = HttpResponse(pdf_data, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="answered_report.pdf"'
    return response


@login_required
def export_unanswered_pdf(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)

    data = _unanswered_dataset(request)
    pdf_data = _draw_table_pdf(
        "Unanswered Report",
        ["time", "caller", "queue", "event", "start_pos", "end_pos", "wait_sec"],
        [[r["time"], r.get("caller", ""), r["queue_display"], r["event"], r["start_pos"], r["end_pos"], r["wait_sec"]] for r in data["rows"]],
    )
    response = HttpResponse(pdf_data, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="unanswered_report.pdf"'
    return response


@login_required
def export_cdr_pdf(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)

    data = _cdr_dataset(request)
    pdf_data = _draw_table_pdf(
        "CDR Report",
        ["uniqueid", "calldate", "operator", "src", "dst", "duration", "billsec", "disp"],
        [
            [
                r.get("uniqueid"),
                r.get("calldate"),
                r.get("operator_display"),
                r.get("src"),
                r.get("dst_display"),
                r.get("duration"),
                r.get("billsec"),
                r.get("disposition"),
            ]
            for r in data["rows"]
        ],
    )
    response = HttpResponse(pdf_data, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="cdr_report.pdf"'
    return response


@login_required
def export_dashboard_traffic_excel(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)
    data = _dashboard_traffic_dataset(request)
    workbook = Workbook()
    daily = workbook.active
    daily.title = "Daily"
    daily.append(["day", "answered", "unanswered", "total"])
    for row in data["daily"]:
        daily.append([row.get("day"), row.get("answered"), row.get("unanswered"), row.get("total")])

    hourly = workbook.create_sheet("Hourly")
    hourly.append(["hour", "answered", "unanswered", "total"])
    for row in data["hourly"]:
        hourly.append([row.get("hour"), row.get("answered"), row.get("unanswered"), row.get("total")])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="dashboard_traffic.xlsx"'
    return response


@login_required
def export_dashboard_queues_excel(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)
    data = _dashboard_queues_dataset(request)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Queues KPI"
    sheet.append(["queue", "total", "answered", "unanswered", "sla", "abandon_rate", "timeout_rate", "avg_hold", "avg_talk", "aht"])
    for row in data["per_queue"]:
        sheet.append(
            [
                row.get("queue_display"),
                row.get("total_calls"),
                row.get("answered"),
                row.get("unanswered"),
                row.get("sla"),
                row.get("abandon_rate"),
                row.get("timeout_rate"),
                row.get("avg_hold"),
                row.get("avg_talk"),
                row.get("aht"),
            ]
        )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="dashboard_queues.xlsx"'
    return response


@login_required
def export_dashboard_operators_excel(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)
    data = _dashboard_operators_dataset(request)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Operators KPI"
    sheet.append(["operator", "answered", "talk_min", "avg_talk", "aht", "share_answered", "rate_per_minute", "payout_total"])
    for row in data["rows_operators"]:
        sheet.append(
            [
                row.get("agent_display"),
                row.get("answered"),
                row.get("talk_min"),
                row.get("avg_talk"),
                row.get("aht"),
                row.get("share_answered"),
                row.get("rate_per_minute"),
                row.get("payout_total"),
            ]
        )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="dashboard_operators.xlsx"'
    return response


@login_required
def export_dashboard_traffic_pdf(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)
    data = _dashboard_traffic_dataset(request)
    plots = [
        {
            "type": "line",
            "title": "Daily total",
            "labels": [r.get("day") for r in data["daily"]],
            "values": [r.get("total") for r in data["daily"]],
        },
        {
            "type": "bar",
            "title": "Hourly total",
            "labels": [r.get("hour") for r in data["hourly"]],
            "values": [r.get("total") for r in data["hourly"]],
        },
    ]
    tables = [
        {
            "title": "Traffic daily table",
            "headers": ["day", "answered", "unanswered", "total"],
            "rows": [[r.get("day"), r.get("answered"), r.get("unanswered"), r.get("total")] for r in data["daily"]],
        },
        {
            "title": "Traffic hourly table",
            "headers": ["hour", "answered", "unanswered", "total"],
            "rows": [[r.get("hour"), r.get("answered"), r.get("unanswered"), r.get("total")] for r in data["hourly"]],
        },
    ]
    pdf_data = _draw_plots_pdf(
        "Dashboard Traffic",
        plots,
        tables=tables,
    )
    response = HttpResponse(pdf_data, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="dashboard_traffic.pdf"'
    return response


@login_required
def export_dashboard_queues_pdf(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)
    data = _dashboard_queues_dataset(request)
    plots = [
        {
            "type": "bar",
            "title": "Calls per queue",
            "labels": [r.get("queue_display") for r in data["per_queue"][:16]],
            "values": [r.get("total_calls") for r in data["per_queue"][:16]],
        },
        {
            "type": "bar",
            "title": "SLA per queue",
            "labels": [r.get("queue_display") for r in data["per_queue"][:16]],
            "values": [r.get("sla") for r in data["per_queue"][:16]],
        },
    ]
    tables = [
        {
            "title": "Queues KPI table",
            "headers": ["queue", "total", "answered", "unanswered", "sla", "abandon", "timeout", "avg_hold", "avg_talk", "aht"],
            "rows": [
                [
                    r.get("queue_display"),
                    r.get("total_calls"),
                    r.get("answered"),
                    r.get("unanswered"),
                    r.get("sla"),
                    r.get("abandon_rate"),
                    r.get("timeout_rate"),
                    r.get("avg_hold"),
                    r.get("avg_talk"),
                    r.get("aht"),
                ]
                for r in data["per_queue"]
            ],
        },
    ]
    pdf_data = _draw_plots_pdf(
        "Dashboard Queues",
        plots,
        tables=tables,
    )
    response = HttpResponse(pdf_data, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="dashboard_queues.pdf"'
    return response


@login_required
def export_dashboard_operators_pdf(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)
    data = _dashboard_operators_dataset(request)
    plots = [
        {
            "type": "bar",
            "title": "Talk minutes per operator",
            "labels": [r.get("agent_display") for r in data["rows_operators"][:16]],
            "values": [r.get("talk_min") for r in data["rows_operators"][:16]],
        },
        {
            "type": "bar",
            "title": "Payout per operator",
            "labels": [r.get("agent_display") for r in data["rows_operators"][:16]],
            "values": [r.get("payout_total") for r in data["rows_operators"][:16]],
        },
    ]
    tables = [
        {
            "title": "Operators KPI table",
            "headers": ["operator", "answered", "talk_min", "avg_talk", "aht", "share", "rate", "payout"],
            "rows": [
                [
                    r.get("agent_display"),
                    r.get("answered"),
                    r.get("talk_min"),
                    r.get("avg_talk"),
                    r.get("aht"),
                    r.get("share_answered"),
                    r.get("rate_per_minute"),
                    r.get("payout_total"),
                ]
                for r in data["rows_operators"]
            ],
        },
    ]
    pdf_data = _draw_plots_pdf(
        "Dashboard Operators",
        plots,
        tables=tables,
    )
    response = HttpResponse(pdf_data, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="dashboard_operators.pdf"'
    return response


@login_required
def export_analytics_excel(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)
    data = _analytics_dataset(request)
    workbook = Workbook()
    daily = workbook.active
    daily.title = "Daily"
    daily.append(["day", "answered", "abandoned", "timeout", "unanswered"])
    for row in data["daily"]:
        daily.append([row.get("day"), row.get("answered"), row.get("abandoned"), row.get("timeout"), row.get("unanswered")])

    hourly = workbook.create_sheet("Hourly")
    hourly.append(["hour", "answered", "unanswered"])
    for row in data["hourly"]:
        hourly.append([row.get("hour"), row.get("answered"), row.get("unanswered")])

    queues = workbook.create_sheet("Queues")
    queues.append(["queue", "total", "answered", "unanswered", "sla", "abandon_rate", "timeout_rate", "aht"])
    for row in data["per_queue"]:
        queues.append(
            [
                row.get("queue_display"),
                row.get("total_calls"),
                row.get("answered"),
                row.get("unanswered"),
                row.get("sla"),
                row.get("abandon_rate"),
                row.get("timeout_rate"),
                row.get("aht"),
            ]
        )

    agents = workbook.create_sheet("Agents")
    agents.append(["agent", "answered", "share", "avg_hold", "avg_talk", "aht"])
    for row in data["top_agents"]:
        agents.append(
            [
                row.get("agent_display"),
                row.get("answered"),
                row.get("share_answered"),
                row.get("avg_hold"),
                row.get("avg_talk"),
                row.get("aht"),
            ]
        )

    callers = workbook.create_sheet("FrequentCallers")
    callers.append(["caller", "calls"])
    for row in data.get("top_callers", []):
        callers.append([row.get("caller"), row.get("calls")])

    operators = workbook.create_sheet("OperatorDurations")
    operators.append(
        [
            "operator",
            "handled_calls",
            "talk_sec_total",
            "talk_min_total",
            "incoming_calls",
            "incoming_talk_min",
            "outgoing_calls",
            "outgoing_talk_min",
            "avg_talk_sec",
            "rank_by_calls",
            "rank_by_talk",
        ]
    )
    for row in data.get("operator_duration_rows", []):
        operators.append(
            [
                row.get("agent_display"),
                row.get("handled_calls"),
                row.get("talk_sec_total"),
                row.get("talk_min_total"),
                row.get("incoming_calls"),
                row.get("incoming_talk_min"),
                row.get("outgoing_calls"),
                row.get("outgoing_talk_min"),
                row.get("avg_talk_sec"),
                row.get("rank_by_calls"),
                row.get("rank_by_talk"),
            ]
        )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="analytics_dashboard.xlsx"'
    return response


@login_required
def export_analytics_pdf(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)
    data = _analytics_dataset(request)
    plots = [
        {
            "type": "line",
            "title": "Daily total calls",
            "labels": [r.get("day") for r in data["daily"][:31]],
            "values": [int(r.get("answered") or 0) + int(r.get("unanswered") or 0) for r in data["daily"][:31]],
        },
        {
            "type": "bar",
            "title": "Hourly total calls",
            "labels": [r.get("hour") for r in data["hourly"][:24]],
            "values": [int(r.get("answered") or 0) + int(r.get("unanswered") or 0) for r in data["hourly"][:24]],
        },
        {
            "type": "bar",
            "title": "Calls by queue",
            "labels": [r.get("queue_display") for r in data["per_queue"][:16]],
            "values": [r.get("total_calls") for r in data["per_queue"][:16]],
        },
        {
            "type": "bar",
            "title": "Answered by operator",
            "labels": [r.get("agent_display") for r in data["top_agents"][:16]],
            "values": [r.get("answered") for r in data["top_agents"][:16]],
        },
        {
            "type": "bar",
            "title": "Frequent callers",
            "labels": [r.get("caller") for r in data.get("top_callers", [])[:16]],
            "values": [r.get("calls") for r in data.get("top_callers", [])[:16]],
        },
        {
            "type": "bar",
            "title": "Operator ranking by handled calls",
            "labels": [r.get("agent_display") for r in data.get("operator_rank_by_calls", [])[:16]],
            "values": [r.get("handled_calls") for r in data.get("operator_rank_by_calls", [])[:16]],
        },
        {
            "type": "bar",
            "title": "Operator ranking by talk minutes",
            "labels": [r.get("agent_display") for r in data.get("operator_rank_by_talk", [])[:16]],
            "values": [r.get("talk_min_total") for r in data.get("operator_rank_by_talk", [])[:16]],
        },
    ]
    tables = [
        {
            "title": "Analytics queues table",
            "headers": ["queue", "total", "answered", "unanswered", "sla", "abandon", "timeout", "aht"],
            "rows": [
                [
                    r.get("queue_display"),
                    r.get("total_calls"),
                    r.get("answered"),
                    r.get("unanswered"),
                    r.get("sla"),
                    r.get("abandon_rate"),
                    r.get("timeout_rate"),
                    r.get("aht"),
                ]
                for r in data.get("per_queue", [])
            ],
        },
        {
            "title": "Analytics agents table",
            "headers": ["operator", "answered", "share", "queues", "avg_hold", "avg_talk", "aht", "talk_min"],
            "rows": [
                [
                    r.get("agent_display"),
                    r.get("answered"),
                    r.get("share_answered"),
                    r.get("queues_count"),
                    r.get("avg_hold"),
                    r.get("avg_talk"),
                    r.get("aht"),
                    r.get("talk_min"),
                ]
                for r in data.get("top_agents", [])
            ],
        },
        {
            "title": "Frequent callers table",
            "headers": ["caller", "calls"],
            "rows": [[r.get("caller"), r.get("calls")] for r in data.get("top_callers", [])],
        },
        {
            "title": "Operator duration table",
            "headers": ["operator", "in_calls", "in_min", "out_calls", "out_min", "total_calls", "total_min", "avg_sec"],
            "rows": [
                [
                    r.get("agent_display"),
                    r.get("incoming_calls"),
                    r.get("incoming_talk_min"),
                    r.get("outgoing_calls"),
                    r.get("outgoing_talk_min"),
                    r.get("handled_calls"),
                    r.get("talk_min_total"),
                    r.get("avg_talk_sec"),
                ]
                for r in data.get("operator_duration_rows", [])
            ],
        },
    ]
    pdf_data = _draw_plots_pdf("Analytics Dashboard", plots, tables=tables)
    response = HttpResponse(pdf_data, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="analytics_dashboard.pdf"'
    return response


@login_required
def realtime_oob_partial(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)
    return render(request, "stats/partials/realtime_oob.html", _build_ami_snapshot(request))
