from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List

from django.http import HttpRequest, HttpResponse
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook

from .datasets import (
    answered_dataset,
    analytics_dataset,
    cdr_dataset,
    dashboard_operators_dataset,
    dashboard_queues_dataset,
    dashboard_traffic_dataset,
    unanswered_dataset,
)
from .helpers import _user_allowed
from .pdf_reports import draw_plots_pdf, draw_table_pdf


@login_required
def export_answered_excel(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)

    data = answered_dataset(request)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Answered"
    sheet.append(["time", "caller", "queue", "agent", "hold_sec", "talk_sec"])
    for row in data["rows"]:
        sheet.append([row["time"], row.get("caller", ""), row["queue_display"], row["agent_display"], row["hold"], row["talk"]])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="answered_report.xlsx"'
    return response


@login_required
def export_unanswered_excel(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)

    data = unanswered_dataset(request)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Unanswered"
    sheet.append(["time", "caller", "queue", "event", "start_pos", "end_pos", "wait_sec"])
    for row in data["rows"]:
        sheet.append(
            [
                row["time"],
                row.get("caller", ""),
                row["queue_display"],
                row["event"],
                row["start_pos"],
                row["end_pos"],
                row["wait_sec"],
            ]
        )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="unanswered_report.xlsx"'
    return response


@login_required
def export_cdr_excel(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)

    data = cdr_dataset(request)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "CDR"
    sheet.append(["uniqueid", "calldate", "operator", "src", "dst", "duration", "billsec", "disposition", "recordingfile"])
    for row in data["rows"]:
        sheet.append(
            [
                row.get("uniqueid"),
                row.get("calldate"),
                row.get("operator_display"),
                row.get("src"),
                row.get("dst_display"),
                row.get("duration"),
                row.get("billsec"),
                row.get("disposition"),
                row.get("recordingfile"),
            ]
        )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="cdr_report.xlsx"'
    return response


@login_required
def export_answered_pdf(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)

    data = answered_dataset(request)
    pdf_data = draw_table_pdf(
        "Answered Report",
        ["time", "caller", "queue", "agent", "hold_sec", "talk_sec"],
        [[r["time"], r.get("caller", ""), r["queue_display"], r["agent_display"], r["hold"], r["talk"]] for r in data["rows"]],
    )
    response = HttpResponse(pdf_data, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="answered_report.pdf"'
    return response


@login_required
def export_unanswered_pdf(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)

    data = unanswered_dataset(request)
    pdf_data = draw_table_pdf(
        "Unanswered Report",
        ["time", "caller", "queue", "event", "start_pos", "end_pos", "wait_sec"],
        [[r["time"], r.get("caller", ""), r["queue_display"], r["event"], r["start_pos"], r["end_pos"], r["wait_sec"]] for r in data["rows"]],
    )
    response = HttpResponse(pdf_data, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="unanswered_report.pdf"'
    return response


@login_required
def export_cdr_pdf(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)

    data = cdr_dataset(request)
    pdf_data = draw_table_pdf(
        "CDR Report",
        ["uniqueid", "calldate", "operator", "src", "dst", "duration", "billsec", "disp"],
        [
            [
                r.get("uniqueid"),
                r.get("calldate"),
                r.get("operator_display"),
                r.get("src"),
                r.get("dst_display"),
                r.get("duration"),
                r.get("billsec"),
                r.get("disposition"),
            ]
            for r in data["rows"]
        ],
    )
    response = HttpResponse(pdf_data, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="cdr_report.pdf"'
    return response


@login_required
def export_dashboard_traffic_excel(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)
    data = dashboard_traffic_dataset(request)
    workbook = Workbook()
    daily = workbook.active
    daily.title = "Daily"
    daily.append(["day", "answered", "unanswered", "total"])
    for row in data["daily"]:
        daily.append([row.get("day"), row.get("answered"), row.get("unanswered"), row.get("total")])

    hourly = workbook.create_sheet("Hourly")
    hourly.append(["hour", "answered", "unanswered", "total"])
    for row in data["hourly"]:
        hourly.append([row.get("hour"), row.get("answered"), row.get("unanswered"), row.get("total")])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="dashboard_traffic.xlsx"'
    return response


@login_required
def export_dashboard_queues_excel(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)
    data = dashboard_queues_dataset(request)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Queues KPI"
    sheet.append(["queue", "total", "answered", "unanswered", "sla", "abandon_rate", "timeout_rate", "avg_hold", "avg_talk", "aht"])
    for row in data["per_queue"]:
        sheet.append(
            [
                row.get("queue_display"),
                row.get("total_calls"),
                row.get("answered"),
                row.get("unanswered"),
                row.get("sla"),
                row.get("abandon_rate"),
                row.get("timeout_rate"),
                row.get("avg_hold"),
                row.get("avg_talk"),
                row.get("aht"),
            ]
        )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="dashboard_queues.xlsx"'
    return response


@login_required
def export_dashboard_operators_excel(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)
    data = dashboard_operators_dataset(request)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Operators KPI"
    sheet.append(["operator", "answered", "talk_min", "avg_talk", "aht", "share_answered", "rate_per_minute", "payout_total"])
    for row in data["rows_operators"]:
        sheet.append(
            [
                row.get("agent_display"),
                row.get("answered"),
                row.get("talk_min"),
                row.get("avg_talk"),
                row.get("aht"),
                row.get("share_answered"),
                row.get("rate_per_minute"),
                row.get("payout_total"),
            ]
        )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="dashboard_operators.xlsx"'
    return response


@login_required
def export_dashboard_traffic_pdf(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)
    data = dashboard_traffic_dataset(request)
    plots = [
        {
            "type": "line",
            "title": "Daily total",
            "labels": [r.get("day") for r in data["daily"]],
            "values": [r.get("total") for r in data["daily"]],
        },
        {
            "type": "bar",
            "title": "Hourly total",
            "labels": [r.get("hour") for r in data["hourly"]],
            "values": [r.get("total") for r in data["hourly"]],
        },
    ]
    tables = [
        {
            "title": "Traffic daily table",
            "headers": ["day", "answered", "unanswered", "total"],
            "rows": [[r.get("day"), r.get("answered"), r.get("unanswered"), r.get("total")] for r in data["daily"]],
        },
        {
            "title": "Traffic hourly table",
            "headers": ["hour", "answered", "unanswered", "total"],
            "rows": [[r.get("hour"), r.get("answered"), r.get("unanswered"), r.get("total")] for r in data["hourly"]],
        },
    ]
    pdf_data = draw_plots_pdf(
        "Dashboard Traffic",
        plots,
        tables=tables,
    )
    response = HttpResponse(pdf_data, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="dashboard_traffic.pdf"'
    return response


@login_required
def export_dashboard_queues_pdf(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)
    data = dashboard_queues_dataset(request)
    plots = [
        {
            "type": "bar",
            "title": "Calls per queue",
            "labels": [r.get("queue_display") for r in data["per_queue"][:16]],
            "values": [r.get("total_calls") for r in data["per_queue"][:16]],
        },
        {
            "type": "bar",
            "title": "SLA per queue",
            "labels": [r.get("queue_display") for r in data["per_queue"][:16]],
            "values": [r.get("sla") for r in data["per_queue"][:16]],
        },
    ]
    tables = [
        {
            "title": "Queues KPI table",
            "headers": ["queue", "total", "answered", "unanswered", "sla", "abandon", "timeout", "avg_hold", "avg_talk", "aht"],
            "rows": [
                [
                    r.get("queue_display"),
                    r.get("total_calls"),
                    r.get("answered"),
                    r.get("unanswered"),
                    r.get("sla"),
                    r.get("abandon_rate"),
                    r.get("timeout_rate"),
                    r.get("avg_hold"),
                    r.get("avg_talk"),
                    r.get("aht"),
                ]
                for r in data["per_queue"]
            ],
        },
    ]
    pdf_data = draw_plots_pdf(
        "Dashboard Queues",
        plots,
        tables=tables,
    )
    response = HttpResponse(pdf_data, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="dashboard_queues.pdf"'
    return response


@login_required
def export_dashboard_operators_pdf(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)
    data = dashboard_operators_dataset(request)
    plots = [
        {
            "type": "bar",
            "title": "Talk minutes per operator",
            "labels": [r.get("agent_display") for r in data["rows_operators"][:16]],
            "values": [r.get("talk_min") for r in data["rows_operators"][:16]],
        },
        {
            "type": "bar",
            "title": "Payout per operator",
            "labels": [r.get("agent_display") for r in data["rows_operators"][:16]],
            "values": [r.get("payout_total") for r in data["rows_operators"][:16]],
        },
    ]
    tables = [
        {
            "title": "Operators KPI table",
            "headers": ["operator", "answered", "talk_min", "avg_talk", "aht", "share", "rate", "payout"],
            "rows": [
                [
                    r.get("agent_display"),
                    r.get("answered"),
                    r.get("talk_min"),
                    r.get("avg_talk"),
                    r.get("aht"),
                    r.get("share_answered"),
                    r.get("rate_per_minute"),
                    r.get("payout_total"),
                ]
                for r in data["rows_operators"]
            ],
        },
    ]
    pdf_data = draw_plots_pdf(
        "Dashboard Operators",
        plots,
        tables=tables,
    )
    response = HttpResponse(pdf_data, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="dashboard_operators.pdf"'
    return response


@login_required
def export_analytics_excel(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)
    data = analytics_dataset(request)
    workbook = Workbook()
    daily = workbook.active
    daily.title = "Daily"
    daily.append(["day", "answered", "abandoned", "timeout", "unanswered"])
    for row in data["daily"]:
        daily.append([row.get("day"), row.get("answered"), row.get("abandoned"), row.get("timeout"), row.get("unanswered")])

    hourly = workbook.create_sheet("Hourly")
    hourly.append(["hour", "answered", "unanswered"])
    for row in data["hourly"]:
        hourly.append([row.get("hour"), row.get("answered"), row.get("unanswered")])

    queues = workbook.create_sheet("Queues")
    queues.append(["queue", "total", "answered", "unanswered", "sla", "abandon_rate", "timeout_rate", "aht"])
    for row in data["per_queue"]:
        queues.append(
            [
                row.get("queue_display"),
                row.get("total_calls"),
                row.get("answered"),
                row.get("unanswered"),
                row.get("sla"),
                row.get("abandon_rate"),
                row.get("timeout_rate"),
                row.get("aht"),
            ]
        )

    agents = workbook.create_sheet("Agents")
    agents.append(["agent", "answered", "share", "avg_hold", "avg_talk", "aht"])
    for row in data["top_agents"]:
        agents.append(
            [
                row.get("agent_display"),
                row.get("answered"),
                row.get("share_answered"),
                row.get("avg_hold"),
                row.get("avg_talk"),
                row.get("aht"),
            ]
        )

    callers = workbook.create_sheet("FrequentCallers")
    callers.append(["caller", "calls"])
    for row in data.get("top_callers", []):
        callers.append([row.get("caller"), row.get("calls")])

    operators = workbook.create_sheet("OperatorDurations")
    operators.append(
        [
            "operator",
            "handled_calls",
            "talk_sec_total",
            "talk_min_total",
            "incoming_calls",
            "incoming_talk_min",
            "outgoing_calls",
            "outgoing_talk_min",
            "avg_talk_sec",
            "rank_by_calls",
            "rank_by_talk",
        ]
    )
    for row in data.get("operator_duration_rows", []):
        operators.append(
            [
                row.get("agent_display"),
                row.get("handled_calls"),
                row.get("talk_sec_total"),
                row.get("talk_min_total"),
                row.get("incoming_calls"),
                row.get("incoming_talk_min"),
                row.get("outgoing_calls"),
                row.get("outgoing_talk_min"),
                row.get("avg_talk_sec"),
                row.get("rank_by_calls"),
                row.get("rank_by_talk"),
            ]
        )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="analytics_dashboard.xlsx"'
    return response


@login_required
def export_analytics_pdf(request: HttpRequest) -> HttpResponse:
    if not _user_allowed(request):
        return HttpResponse("forbidden", status=403)
    data = analytics_dataset(request)
    plots = [
        {
            "type": "line",
            "title": "Daily total calls",
            "labels": [r.get("day") for r in data["daily"][:31]],
            "values": [int(r.get("answered") or 0) + int(r.get("unanswered") or 0) for r in data["daily"][:31]],
        },
        {
            "type": "bar",
            "title": "Hourly total calls",
            "labels": [r.get("hour") for r in data["hourly"][:24]],
            "values": [int(r.get("answered") or 0) + int(r.get("unanswered") or 0) for r in data["hourly"][:24]],
        },
        {
            "type": "bar",
            "title": "Calls by queue",
            "labels": [r.get("queue_display") for r in data["per_queue"][:16]],
            "values": [r.get("total_calls") for r in data["per_queue"][:16]],
        },
        {
            "type": "bar",
            "title": "Answered by operator",
            "labels": [r.get("agent_display") for r in data["top_agents"][:16]],
            "values": [r.get("answered") for r in data["top_agents"][:16]],
        },
        {
            "type": "bar",
            "title": "Frequent callers",
            "labels": [r.get("caller") for r in data.get("top_callers", [])[:16]],
            "values": [r.get("calls") for r in data.get("top_callers", [])[:16]],
        },
        {
            "type": "bar",
            "title": "Operator ranking by handled calls",
            "labels": [r.get("agent_display") for r in data.get("operator_rank_by_calls", [])[:16]],
            "values": [r.get("handled_calls") for r in data.get("operator_rank_by_calls", [])[:16]],
        },
        {
            "type": "bar",
            "title": "Operator ranking by talk minutes",
            "labels": [r.get("agent_display") for r in data.get("operator_rank_by_talk", [])[:16]],
            "values": [r.get("talk_min_total") for r in data.get("operator_rank_by_talk", [])[:16]],
        },
    ]
    pdf_data = draw_plots_pdf(
        "Analytics Dashboard",
        plots,
    )
    response = HttpResponse(pdf_data, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="analytics_dashboard.pdf"'
    return response

